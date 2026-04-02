#!/usr/bin/env python3
"""
validate_batch.py — Full 93-Ticker Accuracy Validation

Runs the complete enrichment pipeline on all 93 reference tickers,
scores the output field-by-field against the embedded ground truth,
and produces:
  1. Console summary (pass rate %, pass/fail per ticker)
  2. validate_results.xlsx  — full scored output with colour coding
  3. validate_mismatches.xlsx — mismatch-only rows for quick review

Usage:
  python validate_batch.py                        # all 93 tickers
  python validate_batch.py DIBS FEAM ACON         # spot-check subset
  python validate_batch.py --resume               # skip already-done rows
  python validate_batch.py --skip-linkedin        # faster, name-only SalesQL

Pass threshold: 80% (670 / 837 fields)
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
from pathlib import Path
from typing import Optional

# ── Make sure we can import the lookup package ────────────────────────────────
APP_DIR = Path(__file__).parent
sys.path.insert(0, str(APP_DIR))

try:
    from dotenv import load_dotenv
    load_dotenv(APP_DIR / ".env")
except ImportError:
    pass

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: pandas and openpyxl are required.  Run: pip install pandas openpyxl")
    sys.exit(1)

from lookup.ticker_resolver  import resolve_ticker, split_name
from lookup.linkedin_finder  import find_linkedin_url
from lookup.salesql_enricher import (
    search_by_name_with_variations, enrich_by_url, _empty as salesql_empty,
)
from lookup.financial_fetcher import fetch_financials_safe
from lookup.ir_finder         import find_ir_data
from lookup.schema_builder    import build_row, empty_row, COLUMN_ORDER
from lookup.email_pattern     import fill_missing_emails


# ─────────────────────────────────────────────────────────────────────────────
# ── Ground Truth ──────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────

GROUND_TRUTH: dict[str, dict] = {
  "DIBS": {"CEO": "David S. Rosenblatt","CFO": "Thomas J. Etergino","CEO EMAIL": "drosenblatt@1stdibs.com","CEO NUMBER": "work +1 646-779-0768","CFO EMAIL": "","CFO NUMBER": "+1 917-379-1470","IR Email": "investors@1stdibs.com","IR Contact": "","IR Page": "https://investors.1stdibs.com/"},
  "FEAM": {"CEO": "Paul Weibel","CFO": "Joshua Malm","CEO EMAIL": "paul.weibel3@gmail.com (no work provided)","CEO NUMBER": "1 610-704-1388","CFO EMAIL": "joshua.malm@gmail.com (no work provided)","CFO NUMBER": "+1 303-875-7548","IR Email": "Not found","IR Contact": "","IR Page": "https://investors.5eadvancedmaterials.com/contact-ir"},
  "AKA": {"CEO": "Ciaran Long","CFO": "Kevin Grant","CEO EMAIL": "ciaran.long@aka-brands.com","CEO NUMBER": "Not found","CFO EMAIL": "kevin.grant@aka-brands.com","CFO NUMBER": "Not found","IR Email": "investors@aka-brands.com","IR Contact": "","IR Page": ""},
  "ACON": {"CEO": "Brent Ness","CFO": "Gregory A. Gould","CEO EMAIL": "bness@aclarion.com","CEO NUMBER": "+1 303-883-4036","CFO EMAIL": "ggould@aclarion.com","CFO NUMBER": "work +1 650-241-1727","IR Email": "ksmith@pcgadvisory.com","IR Contact": "Kirin M. Smith (PCG Advisory)","IR Page": "https://investors.aclarion.com/"},
  "ATGN": {"CEO": "Jeremiah J. Fleming","CFO": "Carolyn David","CEO EMAIL": "Not on Linkedin","CEO NUMBER": "Not on Linkedin","CFO EMAIL": "Not on Linkedin","CFO NUMBER": "Not on Linkedin","IR Email": "ir@altigen.com","IR Contact": "","IR Page": "https://altigen.com/investor-relations/"},
  "AUUA-CA": {"CEO": "Sage Berryman","CFO": "Dale Graham","CEO EMAIL": "sage.berryman@aluula.com","CEO NUMBER": "250-661-0657","CFO EMAIL": "dale.graham@aluula.com","CFO NUMBER": "Not found","IR Email": "IR@aluula.com","IR Contact": "","IR Page": "https://aluula.com/investor-relations/"},
  "ARG-CA": {"CEO": "Aurora G. Davidson","CFO": "Carmen Amezquita Hernandez","CEO EMAIL": "ad@amerigoresources.com","CEO NUMBER": "Not found","CFO EMAIL": "Not on Linkedin","CFO NUMBER": "Not on Linkedin","IR Email": "info@amerigoresources.com","IR Contact": "Aurora Davidson, CEO","IR Page": "https://www.amerigoresources.com/company/contact/"},
  "ASYS": {"CEO": "Robert C. Daigle","CFO": "Mark D. Weaver","CEO EMAIL": "bdaigle@amtechsystems.com","CEO NUMBER": "work +1 480-967-5146","CFO EMAIL": "Not found","CFO NUMBER": "Not found","IR Email": "irelations@amtechsystems.com","IR Contact": "Mark Weaver, Interim CFO","IR Page": "https://www.amtechsystems.com/investors"},
  "AMNF": {"CEO": "Deanna Jurgens","CFO": "Andrew Leonard","CEO EMAIL": "djurgens@armaninofoods.com","CEO NUMBER": "work +1 510-441-9011","CFO EMAIL": "aleonard@armaninofoods.com","CFO NUMBER": "work +1 510-441-9011","IR Email": "Not found","IR Contact": "","IR Page": "https://amerityre.com/investor-relations/"},
  "ACNT": {"CEO": "J. Bryan Kitchen","CFO": "Ryan Kavalauskas","CEO EMAIL": "jkitchen@AscentCo.com","CEO NUMBER": "work +1 630-884-9181","CFO EMAIL": "rkavalauskas@ascentco.com","CFO NUMBER": "work +1 630-884-9181","IR Email": "investorrelations@ascentco.com","IR Contact": "","IR Page": "https://ir.ascentco.com/"},
  "SALT": {"CEO": "","CFO": "","CEO EMAIL": "npeterson@atlassalt.com","CEO NUMBER": "778-995-4619","CFO EMAIL": "jkilborn@atlassalt.com","CFO NUMBER": "work +1 709-275-2009","IR Email": "Not found","IR Contact": "","IR Page": "https://www.saltxtechnology.com/investor/"},
  "AYTU": {"CEO": "Joshua R. Disbrow","CFO": "Ryan J. Selhorn","CEO EMAIL": "jdisbrow@aytubio.com","CEO NUMBER": "work +1 720-437-6580","CFO EMAIL": "rselhorn@aytubio.com","CFO NUMBER": "work +1 720-437-6580","IR Email": "investors@aytubio.com","IR Contact": "","IR Page": "https://investors.aytubio.com/"},
  "BLZE": {"CEO": "Gleb Budman","CFO": "Marc Suidan","CEO EMAIL": "budmang@backblaze.com","CEO NUMBER": "work +1 650-352-3738","CFO EMAIL": "marc@backblaze.com","CFO NUMBER": "408-607-5852","IR Email": "IR@backblaze.com","IR Contact": "","IR Page": "https://ir.backblaze.com/"},
  "BYL": {"CEO": "","CFO": "","CEO EMAIL": "leighton.carroll@baylintech.com","CEO NUMBER": "+1 201-888-8428","CFO EMAIL": "cliff.gary@baylintech.com","CFO NUMBER": "work +1 416-222-1888","IR Email": "kelly.myles@baylintech.com","IR Contact": "Kelly Myles, Director IR","IR Page": "https://www.baylintech.com/shareholder-information/"},
  "RX-CA": {"CEO": "René C. Goehrum","CFO": "Robert March","CEO EMAIL": "rgoehrum@biosyent.com","CEO NUMBER": "N/a","CFO EMAIL": "rmarch@biosyent.com","CFO NUMBER": "work +1 905-206-0013","IR Email": "investors@biosyent.com","IR Contact": "","IR Page": "https://www.biosyent.com/investors/"},
  "BLM": {"CEO": "","CFO": "","CEO EMAIL": "smacfabe@blumetric.ca","CEO NUMBER": "work +1 877-487-8436","CFO EMAIL": "dhilton@blumetric.ca","CFO NUMBER": "+1 877-487-8436","IR Email": "smacfabe@blumetric.ca","IR Contact": "Scott MacFabe, CEO","IR Page": "https://www.blumetric.ca/investor-relations/"},
  "BOGO-CA": {"CEO": "Kelly Malcolm","CFO": "Lisanna Lewis","CEO EMAIL": "Not on Linkedin","CEO NUMBER": "Not on Linkedin","CFO EMAIL": "lisanna.lewis@elkomininggroup.com","CFO NUMBER": "","IR Email": "kmalcolm@borealismining.com","IR Contact": "Kelly Malcolm, CEO & Director","IR Page": "https://borealismining.com/investors/"},
  "BOF": {"CEO": "Eric Healy","CFO": "John Dalfonsi","CEO EMAIL": "eric@branchoutfood.com","CEO NUMBER": "541-419-4078","CFO EMAIL": "Not found","CFO NUMBER": "Not found","IR Email": "ir@branchoutfood.com","IR Contact": "","IR Page": "https://branchoutfood.com/pages/investor-relations"},
  "BUDA": {"CEO": "Horatio Lonsdale-Hands","CFO": "Clint Bowers","CEO EMAIL": "horatio@budajuice.com","CEO NUMBER": "work +1 214-308-5003","CFO EMAIL": "","CFO NUMBER": "","IR Email": "brian@haydenir.com","IR Contact": "Brian S. Siegel (Hayden IR)","IR Page": "https://budajuice.com/investors"},
  "BILD.V": {"CEO": "","CFO": "","CEO EMAIL": "shawnwilson@builddirect.com","CEO NUMBER": "714-333-7848","CFO EMAIL": "kerrybiggs@builddirect.com","CFO NUMBER": "work +1 604-662-8100","IR Email": "ir@builddirect.com","IR Contact": "","IR Page": "https://ir.builddirect.com/"},
  "CWD": {"CEO": "John Christopher Loeffler","CFO": "Jade Leung","CEO EMAIL": "chris@caliberco.com","CEO NUMBER": "+1 480-295-7600","CFO EMAIL": "jade.leung@caliberco.com","CFO NUMBER": "602-329-7682","IR Email": "ir@caliberco.com","IR Contact": "Ilya Grozovsky, VP IR","IR Page": "https://ir.caliberco.com"},
  "LOVFF": {"CEO": "Zohar Krivorot","CFO": "Nicholas Sosiak","CEO EMAIL": "zohar@cannara.ca","CEO NUMBER": "514-830-0788","CFO EMAIL": "nicholas.sosiak@cannara.ca","CFO NUMBER": "work +1 514-543-4200","IR Email": "investors@cannara.ca","IR Contact": "","IR Page": "https://www.cannara.ca"},
  "CITR": {"CEO": "Wesley J. Bolsen","CFO": "Nanuk Warman","CEO EMAIL": "","CEO NUMBER": "650-387-9962","CFO EMAIL": "nanuk@pubcoreporting.com","CFO NUMBER": "work +1 909-519-5470","IR Email": "CITR@haydenir.com","IR Contact": "Brett Maas (Hayden IR)","IR Page": "https://ir.citrotech.com/"},
  "LODE": {"CEO": "Corrado F. de Gasperis","CFO": "Judd B. Merrill","CEO EMAIL": "degasperis@comstock.inc","CEO NUMBER": "+1 775-847-4755","CFO EMAIL": "","CFO NUMBER": "","IR Email": "ir@comstock.inc","IR Contact": "Judd B. Merrill, CFO","IR Page": "https://comstock.inc/investors/contact-us/"},
  "CREX": {"CEO": "Richard C. Mills","CFO": "Tamra L. Koshewa","CEO EMAIL": "rick.mills@cri.com","CEO NUMBER": "work +1 888-323-3633","CFO EMAIL": "tamra.koshewa@cri.com","CFO NUMBER": "work +1 888-323-3633","IR Email": "Not found","IR Contact": "","IR Page": "https://investors.cri.com/"},
  "CXDO": {"CEO": "Jeffrey G. Korn","CFO": "Ronald Vincent","CEO EMAIL": "jkorn@crexendo.com","CEO NUMBER": "work +1 602-714-8500","CFO EMAIL": "N/a","CFO NUMBER": "","IR Email": "ir@crexendo.com","IR Contact": "","IR Page": "https://www.crexendo.com/for-investors/"},
  "CTW": {"CEO": "Ryuichi Sasaki","CFO": "Patrick Liu","CEO EMAIL": "Not found on Linkedin","CEO NUMBER": "Not found on Linkedin","CFO EMAIL": "Not found on Linkedin","CFO NUMBER": "Not found on Linkedin","IR Email": "investor@ctw.inc","IR Contact": "","IR Page": "https://ctw.inc/investors"},
  "CYBT.CN": {"CEO": "","CFO": "","CEO EMAIL": "justin@cybeats.com","CEO NUMBER": "work +1 888-832-9232","CFO EMAIL": "josh@cybeats.com","CFO NUMBER": "work +1 888-832-9232","IR Email": "ir@cybeats.com","IR Contact": "James Van Staveren","IR Page": "https://www.cybeats.com/"},
  "DAIO": {"CEO": "William Wentworth","CFO": "Charles J. DiBona","CEO EMAIL": "","CEO NUMBER": "work +1 425-867-6922","CFO EMAIL": "dibonac@dataio.com","CFO NUMBER": "+1 425-867-6922","IR Email": "","IR Contact": "","IR Page": "https://www.dataio.com/Contact-Us/Investor-Relations"},
  "DHX": {"CEO": "Art Zeile","CFO": "Gregory Schippers","CEO EMAIL": "art.zeile@dhigroupinc.com","CEO NUMBER": "","CFO EMAIL": "greg.schippers@dhigroupinc.com","CFO NUMBER": "work +1 212-725-6550","IR Email": "ir@dhigroupinc.com","IR Contact": "Todd Kehrli / Jim Byers","IR Page": "https://dhigroupinc.com/investors/"},
  "DRT-CA": {"CEO": "Benjamin Urban","CFO": "Fareeha Khan","CEO EMAIL": "burban@dirtt.com","CEO NUMBER": "work +1 212-725-6550","CFO EMAIL": "","CFO NUMBER": "work +1 800-605-6707","IR Email": "ir@dirtt.com","IR Contact": "","IR Page": "https://www.dirtt.com/investors/"},
  "DCGO": {"CEO": "Bienstock Lee","CFO": "Norman Rosenberg","CEO EMAIL": "lee@docgo.com","CEO NUMBER": "Not found","CFO EMAIL": "nrosenberg@docgo.com","CFO NUMBER": "Not found","IR Email": "ir@docgo.com","IR Contact": "","IR Page": "https://ir.docgo.com/"},
  "DLPN": {"CEO": "William O'Dowd","CFO": "Mirta A. Negrini","CEO EMAIL": "bill@dolphinentertainment.com","CEO NUMBER": "work +1 305-774-0407","CFO EMAIL": "mirta@dolphinentertainment.com","CFO NUMBER": "work +1 305-774-0407","IR Email": "Not found","IR Contact": "","IR Page": "https://dolphinentertainment.com/investor-relations/"},
  "DAIC": {"CEO": "Edmund Nabrotzky","CFO": "Charlie Maddox","CEO EMAIL": "ed.nabrotzky@seeidinc.com","CEO NUMBER": "work +1 888-733-4301","CFO EMAIL": "charlie@seeidinc.com","CFO NUMBER": "+1 702-518-0899","IR Email": "dotai@icrinc.com","IR Contact": "","IR Page": "https://ir.daic.ai/"},
  "ETST": {"CEO": "Giorgio R. Saumat","CFO": "Ernesto L. Flores","CEO EMAIL": "","CEO NUMBER": "","CFO EMAIL": "","CFO NUMBER": "+1 786-312-3588","IR Email": "Not found","IR Contact": "","IR Page": ""},
  "ELWT": {"CEO": "Barry Rubens","CFO": "Sean Arnette","CEO EMAIL": "brubens@elauwit.com","CEO NUMBER": "+1 704-662-2975","CFO EMAIL": "sean@elauwit.co","CFO NUMBER": "work +1 704-558-3099","IR Email": "mkreps@darrowir.com","IR Contact": "Matt Kreps (Darrow IR)","IR Page": "https://investors.elauwit.com/"},
  "ECOR": {"CEO": "Daniel S. Goldberger","CFO": "Joshua S. Lev","CEO EMAIL": "dan.goldberger@electrocore.com","CEO NUMBER": "work +1 888-903-2673","CFO EMAIL": "joshua.lev@electrocore.com","CFO NUMBER": "work +1 888-903-2673","IR Email": "Investors@electrocore.com","IR Contact": "","IR Page": "https://investor.electrocore.com/"},
  "WATT": {"CEO": "Mallorie Sara Burak","CFO": "Mallorie Sara Burak","CEO EMAIL": "mburak@energous.com","CEO NUMBER": "+1 408-893-9311","CFO EMAIL": "mburak@energous.com","CFO NUMBER": "+1 408-893-9311","IR Email": "ir@energous.com","IR Contact": "","IR Page": "https://ir.energous.com/"},
  "COCH": {"CEO": "Brent T. Lucas","CFO": "Robert Potashnick","CEO EMAIL": "blucas@envoymedical.com","CEO NUMBER": "work +1 651-361-8000","CFO EMAIL": "","CFO NUMBER": "","IR Email": "InvestorRelations@EnvoyMedical.com","IR Contact": "Phil Carlson (KCSA)","IR Page": "https://ir.envoymedical.com/"},
  "ENW-CA": {"CEO": "Brent Charleton","CFO": "Dylan Murray","CEO EMAIL": "bcharleton@enwave.net","CEO NUMBER": "work +1 604-806-6110","CFO EMAIL": "dmurray@enwave.net","CFO NUMBER": "work +1 604-806-6110","IR Email": "info@enwave.net","IR Contact": "Noel Atkinson","IR Page": "https://www.enwave.net/investors/"},
  "EONR": {"CEO": "Dante V. Caravaggio","CFO": "Mitchell B. Trotter","CEO EMAIL": "dante@swiexcavating.com","CEO NUMBER": "","CFO EMAIL": "","CFO NUMBER": "","IR Email": "Mike@PLRinvest.com","IR Contact": "Mike Porter","IR Page": "https://www.eon-r.com/investor-relations"},
  "FCI": {"CEO": "Mark St. Hill","CFO": "Carl Lewis","CEO EMAIL": "Not found on Linkedin","CEO NUMBER": "Not found on Linkedin","CFO EMAIL": "Not found on Linkedin","CFO NUMBER": "Not found on Linkedin","IR Email": "InvestorRelations@cfindustries.com","IR Contact": "","IR Page": ""},
  "FTK": {"CEO": "Ryan G. Ezell","CFO": "James Bond Clement","CEO EMAIL": "rezell@flotekind.com","CEO NUMBER": "work +1 800-256-4703","CFO EMAIL": "Not on SQL","CFO NUMBER": "Not on SQL","IR Email": "ir@flotekind.com","IR Contact": "Mike Critelli","IR Page": "https://ir.flotekind.com/"},
  "FOR-CA": {"CEO": "Dale Verran","CFO": "Patrick John McGrath","CEO EMAIL": "dverran@fortunebaycorp.com","CEO NUMBER": "work +1 902-422-1421","CFO EMAIL": "","CFO NUMBER": "work +1 902-422-1421","IR Email": "investor@forrester.com","IR Contact": "","IR Page": ""},
  "FET": {"CEO": "Neal A. Lux","CFO": "D. Lyle Williams","CEO EMAIL": "neal.lux@f-e-t.com","CEO NUMBER": "work +1 713-351-7900","CFO EMAIL": "lyle.williams@f-e-t.com","CFO NUMBER": "+1 713-412-1324","IR Email": "IR@f-e-t.com","IR Contact": "Rob Kukla","IR Page": "https://ir.f-e-t.com/"},
  "KBSX": {"CEO": "David Chuang","CFO": "Sebastian Tadla","CEO EMAIL": "","CEO NUMBER": "","CFO EMAIL": "","CFO NUMBER": "work +1 303-444-2226","IR Email": "","IR Contact": "Scott Powell (Skyline Corp Comms)","IR Page": ""},
  "GELS": {"CEO": "Nathan J. Givoni","CFO": "Thuy-Linh Gigler","CEO EMAIL": "","CEO NUMBER": "","CFO EMAIL": "Not found","CFO NUMBER": "Not found","IR Email": "ir@gelteq.com","IR Contact": "Matt Kreps (Darrow Associates)","IR Page": "https://ir.gelteq.com/"},
  "HITI": {"CEO": "Raj Grover","CFO": "Mayank Mahajan","CEO EMAIL": "raj@hightideinc.ca","CEO NUMBER": "","CFO EMAIL": "mayank@hightideinc.com","CFO NUMBER": "780-201-9007","IR Email": "ir@hightideinc.com","IR Contact": "","IR Page": "https://hightideinc.com/investor-faq/"},
  "IDR": {"CEO": "John A. Swallow","CFO": "Grant A. Brackebusch","CEO EMAIL": "","CEO NUMBER": "","CFO EMAIL": "Not found","CFO NUMBER": "Not found","IR Email": "Not found","IR Contact": "","IR Page": "https://www.idahostrategic.com/"},
  "IEHC": {"CEO": "David Offerman","CFO": "Subrata Purkayastha","CEO EMAIL": "dave@iehcorp.com","CEO NUMBER": "","CFO EMAIL": "spurkayastha@iehcorp.com","CFO NUMBER": "865-382-1028","IR Email": "","IR Contact": "Thomas L. Barbato, CFO","IR Page": "https://www.iec-electronics.com/investors/overview"},
  "IDN": {"CEO": "Bryan Lewis","CFO": "Adam Sragovicz","CEO EMAIL": "blewis@intellicheck.com","CEO NUMBER": "work +1 516-992-1900","CFO EMAIL": "asragovicz@intellicheck.com","CFO NUMBER": "work +1 516-992-1900","IR Email": "gjackson@intellicheck.com","IR Contact": "Gar Jackson","IR Page": "https://www.intellicheck.com/investors"},
  "IVF": {"CEO": "Steven M. Shum","CFO": "Terah Krigsvold","CEO EMAIL": "steveshum@invobioscience.com","CEO NUMBER": "work +1 978-878-9505","CFO EMAIL": "tkrigsvold@invofertility.com","CFO NUMBER": "","IR Email": "Not found","IR Contact": "","IR Page": "https://invofertility.com/investors/"},
  "IRIX": {"CEO": "Patrick Mercer","CFO": "Romeo R. Dizon","CEO EMAIL": "pmercer@iridex.com","CEO NUMBER": "work +1 650-940-4710","CFO EMAIL": "rdizon@iridex.com","CFO NUMBER": "+1 650-940-4710","IR Email": "investors@iridex.com","IR Contact": "Philip Taylor (Gilmartin Group)","IR Page": "https://iridex.gcs-web.com/"},
  "JRSH": {"CEO": "Lin Hung Choi","CFO": "Gilbert K. Lee","CEO EMAIL": "Not found on Linkedin","CEO NUMBER": "Not found on Linkedin","CFO EMAIL": "","CFO NUMBER": "","IR Email": "Not found","IR Contact": "","IR Page": "https://ir.jerashholdings.com/"},
  "KDOZ-CA": {"CEO": "Jason Miles Williams","CFO": "Henry Walford Bromley","CEO EMAIL": "jason@kidoz.net","CEO NUMBER": "Not found on Linkedin","CFO EMAIL": "Not found","CFO NUMBER": "Not Found","IR Email": "ir@koilenergy.com","IR Contact": "Trevor Ashurst, VP Finance","IR Page": "https://www.koilenergy.com/Investors/"},
  "KLNG": {"CEO": "Erik Wiik","CFO": "Kurt Keller","CEO EMAIL": "ewiik@koilenergy.com","CEO NUMBER": "work +1 281-862-2201","CFO EMAIL": "kkeller@koilenergy.com","CFO NUMBER": "713-857-7560","IR Email": "KOPN@mzgroup.us","IR Contact": "Erich Manz, CFO","IR Page": "https://www.kopin.com/investors/"},
  "KOPN": {"CEO": "Michael Murray","CFO": "Erich Manz","CEO EMAIL": "mmurray@kopin.com","CEO NUMBER": "339-927-0927","CFO EMAIL": "emanz@kopin.com","CFO NUMBER": "508-870-5959","IR Email": "Not found","IR Contact": "","IR Page": "https://www.libertystream.com/"},
  "LIB-CA": {"CEO": "Alexander J. Wylie","CFO": "Morgan Tiernan","CEO EMAIL": "","CEO NUMBER": "N/a","CFO EMAIL": "","CFO NUMBER": "work +1 604-428-9480","IR Email": "MAMA@mzgroup.us","IR Contact": "Lucas A. Zimmerman (MZ Group)","IR Page": "https://ir.mamascreations.com/"},
  "MAMA": {"CEO": "Adam L. Michaels","CFO": "Anthony J. Gruber","CEO EMAIL": "adam@mamamancinis.com","CEO NUMBER": "work +1 201-531-1212","CFO EMAIL": "anthonyg@mamamancinis.com","CFO NUMBER": "201-274-9664","IR Email": "Not found","IR Contact": "","IR Page": "https://www.mccoyglobal.com/investors/"},
  "MCB-CA": {"CEO": "Jim Rakievich","CFO": "Lindsay Marie Mcgill","CEO EMAIL": "jrakievich@mccoyglobal.com","CEO NUMBER": "work +1 780-453-8451","CFO EMAIL": "lmcgill@mccoyglobal.com","CFO NUMBER": "+1 780-453-8451","IR Email": "","IR Contact": "","IR Page": "https://investors.mdb.com/"},
  "MDBH": {"CEO": "Chris Marlett","CFO": "Jeremy W. James","CEO EMAIL": "m@mdb.com","CEO NUMBER": "work +1 945-262-9010","CFO EMAIL": "jjames@mdb.com","CFO NUMBER": "work +1 945-262-9010","IR Email": "IR@ncsmultistage.com","IR Contact": "Mike Morrison, CFO","IR Page": "https://ir.ncsmultistage.com/"},
  "NCSM": {"CEO": "Ryan Hummer","CFO": "Michael L. Morrison","CEO EMAIL": "ryan@ncsmultistage.com","CEO NUMBER": "work +1 281-453-2222","CFO EMAIL": "mmorrison@ncsmultistage.com","CFO NUMBER": "work +1 281-453-2222","IR Email": "browe@newtekone.com","IR Contact": "Bryce Rowe","IR Page": "https://investor.newtekbusinessservices.com/"},
  "NEWT": {"CEO": "Barry Sloane","CFO": "Frank DeMaria","CEO EMAIL": "bsloane@newtekone.com","CEO NUMBER": "work +1 855-763-9835","CFO EMAIL": "fdemaria@newtekone.com","CFO NUMBER": "+1 718-986-6346","IR Email": "NWTG@encore-ir.com","IR Contact": "Ron Both / Grant Stude (Encore IR)","IR Page": "https://www.newtongolfir.com/"},
  "NWTG": {"CEO": "Gregor Campbell","CFO": "Jeffery R. Clayborne","CEO EMAIL": "gcampbell@newtongolfco.com","CEO NUMBER": "","CFO EMAIL": "jclayborne@newtongolfco.com","CFO NUMBER": "","IR Email": "","IR Contact": "","IR Page": "https://nexliving.ca/"},
  "NXLV-CA": {"CEO": "Stavro Stathonikos","CFO": "Glenn A. Holmes","CEO EMAIL": "sstathonikos@nexliving.ca","CEO NUMBER": "Not found","CFO EMAIL": "Not found on Linkedin","CFO NUMBER": "Not found on Linkedin","IR Email": "NTRP@mzgroup.us","IR Contact": "Chris Tyson / Larry Holub (MZ Group)","IR Page": "https://investors.nexttrip.com/contact-ir"},
  "NTRP": {"CEO": "Bill Kerby","CFO": "","CEO EMAIL": "bill.kerby@nexttrip.com","CEO NUMBER": "+1 954-888-9779","CFO EMAIL": "Not found on Linkedin","CFO NUMBER": "Not found on Linkedin","IR Email": "cdobbin@novaleaphealth.com","IR Contact": "Chris Dobbin","IR Page": "https://novaleaphealth.com/"},
  "NLH-CA": {"CEO": "Christopher Dobbin","CFO": "Chris LeBlanc","CEO EMAIL": "cdobbin@novaleaphealth.com","CEO NUMBER": "+1 902-401-9480","CFO EMAIL": "cleblanc@novaleaphealth.com","CFO NUMBER": "work +1 902-401-9480","IR Email": "ir@omsos.com","IR Contact": "Brandi Piacente (Piacente Financial)","IR Page": "https://ir.omsos.com/"},
  "OMSE": {"CEO": "Meng Hock How","CFO": "Yeo Kevin","CEO EMAIL": "menghock.how@omsos.com","CEO NUMBER": "","CFO EMAIL": "kevin.yeo@omsos.com","CFO NUMBER": "work +65 6861 2677","IR Email": "investor@optimumbank.com","IR Contact": "Seth Denison","IR Page": "https://optimumbankholdings.q4ir.com/"},
  "OPHC": {"CEO": "Timothy L. Terry","CFO": "Elliot Nunez","CEO EMAIL": "tterry8757@gmail.com","CEO NUMBER": "+1 561-309-4091","CFO EMAIL": "enunez@optimumbank.com","CFO NUMBER": "work +1 954-900-2800","IR Email": "Not found","IR Contact": "","IR Page": ""},
  "PALS-CA": {"CEO": "Peter James Shippen","CFO": "Sarah Zilik","CEO EMAIL": "peter.shippen@paragongeochem.com","CEO NUMBER": "+1 416-304-6811","CFO EMAIL": "Not found on Linkedin","CFO NUMBER": "Not found on Linkedin","IR Email": "Investor_Relations@PerfectCorp.com","IR Contact": "","IR Page": "https://ir.perfectcorp.com/"},
  "PERF": {"CEO": "Alice H. Chang","CFO": "","CEO EMAIL": "alice@perfectcorp.com","CEO NUMBER": "Not on SQL","CFO EMAIL": "Not found on Linkedin","CFO NUMBER": "Not found on Linkedin","IR Email": "plby@fnkir.com","IR Contact": "","IR Page": "https://www.plbygroup.com/investors/"},
  "PPSI": {"CEO": "Nathan J. Mazurek","CFO": "Walter Michalec","CEO EMAIL": "nathan@pioneerpowersolutions.com","CEO NUMBER": "work +1 212-867-0700","CFO EMAIL": "walter@pioneerpowersolutions.com","CFO NUMBER": "work +1 212-867-0700","IR Email": "investors@precipiodx.com","IR Contact": "","IR Page": "https://www.precipiodx.com/investors/"},
  "PLBY": {"CEO": "Ben Kohn","CFO": "Marc B. Crossman","CEO EMAIL": "bkohn@playboy.com","CEO NUMBER": "work +1 212-692-2000","CFO EMAIL": "Mcrossman@playboy.com","CFO NUMBER": "work +1 310-424-1800","IR Email": "investor.relations@pro-dex.com","IR Contact": "","IR Page": "https://www.pro-dex.com/investors"},
  "PRPO": {"CEO": "Ilan Danieli","CFO": "Matthew Gage","CEO EMAIL": "idanieli@precipiodx.com","CEO NUMBER": "work +1 203-787-7888","CFO EMAIL": "mgage@precipiodx.com","CFO NUMBER": "work +1 203-787-7888","IR Email": "ir@usaqcorp.com","IR Contact": "Olivia Giamanco","IR Page": "https://www.usaqcorp.com/investor-relations"},
  "PDEX": {"CEO": "Richard Lee van Kirk","CFO": "Alisha K. Charlton","CEO EMAIL": "rick.vankirk@pro-dex.com","CEO NUMBER": "work +1 562-502-1100","CFO EMAIL": "alisha.charlton@pro-dex.com","CFO NUMBER": "Not found","IR Email": "ir@qualtekservices.com","IR Contact": "","IR Page": "https://investors.qualtekservices.com/"},
  "USAQ": {"CEO": "Troy Grogan","CFO": "Troy Grogan","CEO EMAIL": "troy@medicalpracticeincome.com","CEO NUMBER": "","CFO EMAIL": "troy@medicalpracticeincome.com","CFO NUMBER": "","IR Email": "","IR Contact": "","IR Page": "https://www.rubiconorganics.com/investors/"},
  "QBAK": {"CEO": "Steven Nathan Bronson","CFO": "Ryan J. Hoffman","CEO EMAIL": "Not on SQL","CEO NUMBER": "Not on SQL","CFO EMAIL": "","CFO NUMBER": "work +1 805-484-8855","IR Email": "","IR Contact": "","IR Page": "https://sanuwave.com/investors/"},
  "ROMJ-CA": {"CEO": "Margaret Brodie","CFO": "Glen W. Ibbott","CEO EMAIL": "margaret.brodie@rubiconorganics.com","CEO NUMBER": "work +1 647-206-1231","CFO EMAIL": "glen.ibbott@rubiconorganics.com","CFO NUMBER": "work +1 604-331-1296","IR Email": "","IR Contact": "Lucas A. Zimmerman (MZ North America)","IR Page": "https://ir.skyplug.com/"},
  "SNWV": {"CEO": "Morgan C. Frank","CFO": "Peter Sorensen","CEO EMAIL": "morgan.frank@sanuwave.com","CEO NUMBER": "work +1 770-419-7525","CFO EMAIL": "peter.sorensen@sanuwave.com","CFO NUMBER": "770-419-7525","IR Email": "investor_relations@sunlife.com","IR Contact": "","IR Page": ""},
  "SKYX": {"CEO": "Leonard Jay Sokolow","CFO": "Marc-Andre Boisseau","CEO EMAIL": "","CEO NUMBER": "+1 561-922-3559","CFO EMAIL": "Not found on Linkedin","CFO NUMBER": "Not found on Linkedin","IR Email": "info@strawberryfieldsreit.com","IR Contact": "Jeffrey Bajtner","IR Page": "https://www.strawberryfieldsreit.com/investors/"},
  "SUNXF": {"CEO": "Mark Tadros","CFO": "Vitaly Melnikov","CEO EMAIL": "mark@stardustsolar.com","CEO NUMBER": "work +1 888-620-6733","CFO EMAIL": "vitaly@stardustsolar.com","CFO NUMBER": "work +1 888-620-6733","IR Email": "Not found","IR Contact": "","IR Page": "https://tantalus.com/investors/"},
  "STRW": {"CEO": "Moishe Gubin","CFO": "Greg Flamion","CEO EMAIL": "mgubin@strawberryfieldsreit.com","CEO NUMBER": "work +1 954-900-2800","CFO EMAIL": "not on","CFO NUMBER": "work +1 574-807-0800","IR Email": "investor.relations@telus.com","IR Contact": "Ian McMillan","IR Page": "https://www.telus.com/en/about/investor-relations/"},
  "GRID-CA": {"CEO": "Peter A. Londa","CFO": "Azim Lalani","CEO EMAIL": "plonda@tantalus.com","CEO NUMBER": "+1 917-593-6574","CFO EMAIL": "alalani@tantalus.com","CFO NUMBER": "work +1 604-299-0458","IR Email": "info@thermalenergy.com","IR Contact": "","IR Page": "https://www.thermalenergy.com/investors.html"},
  "TELI-CA": {"CEO": "Henry Dubina","CFO": "John Kirincic","CEO EMAIL": "henry.dubina@mt.com","CEO NUMBER": "","CFO EMAIL": "Not found on Linkedin","CFO NUMBER": "Not found on Linkedin","IR Email": "info@tssiusa.com","IR Contact": "James Carbonara / Brett Maas (Hayden IR)","IR Page": "https://ir.tssiusa.com/"},
  "TMG-CA": {"CEO": "William M. Crossland","CFO": "Jie Zhang","CEO EMAIL": "bill.crossland@thermalenergy.com","CEO NUMBER": "work +1 613-723-6776","CFO EMAIL": "","CFO NUMBER": "work +1 613-723-6776","IR Email": "investors@unusualmachines.com","IR Contact": "Christine Petraglia (CS Investor Relations)","IR Page": "https://www.unusualmachines.com/"},
  "TSSI": {"CEO": "Darryll E. Dewan","CFO": "Daniel M. Chism","CEO EMAIL": "ddewan@tssiusa.com","CEO NUMBER": "work +1 512-310-1000","CFO EMAIL": "dchism@tssiusa.com","CFO NUMBER": "+1 512-431-1134","IR Email": "paul.manley@usio.com","IR Contact": "Paul Manley","IR Page": "https://usio.com/investor-relations/"},
  "UMAC": {"CEO": "Allan Evans","CFO": "Brian Hoff","CEO EMAIL": "allan@unusualmachines.com","CEO NUMBER": "+1 509-378-4685","CFO EMAIL": "brian@unusualmachines.com","CFO NUMBER": "","IR Email": "investors@velo3d.com","IR Contact": "James Carbonara (Hayden IR)","IR Page": "https://ir.velo3d.com"},
  "USIO": {"CEO": "Louis A. Hoch","CFO": "","CEO EMAIL": "louis.hoch@usio.com","CEO NUMBER": "work +1 210-249-4100","CFO EMAIL": "","CFO NUMBER": "","IR Email": "investors@xtractone.com","IR Contact": "","IR Page": "https://investors.xtractone.com/"},
  "VELO": {"CEO": "Arun Jeldi","CFO": "Bernard Chung","CEO EMAIL": "arun.jeldi@velo3d.com","CEO NUMBER": "","CFO EMAIL": "","CFO NUMBER": "+1 480-433-1661","IR Email": "aladha@zedcor.ca","IR Contact": "","IR Page": "https://zedcor.com/investor-relations/"},
  "XTRA-CA": {"CEO": "Peter Evans","CFO": "Karen Hersh","CEO EMAIL": "peter.evans@optiv.com","CEO NUMBER": "+1 613-533-2000","CFO EMAIL": "karenh@patriot1tech.com","CFO NUMBER": "+1 888-728-1832","IR Email": "Not found","IR Contact": "","IR Page": "https://investor.zedge.net/"},
  "ZDC-CA": {"CEO": "Todd Michael Ziniuk","CFO": "Amin Ladha","CEO EMAIL": "bubba@zedcor.ca","CEO NUMBER": "+1 780-612-8644","CFO EMAIL": "aladha@zedcor.ca","CFO NUMBER": "+1 780-612-8644","IR Email": "ir@zoomd.com","IR Contact": "Amit Bohensky, Chairman","IR Page": "https://zoomd.com/investors/"},
  "ZDGE": {"CEO": "Jonathan Reich","CFO": "Yi Tsai","CEO EMAIL": "jonathan.reich@zedge.net","CEO NUMBER": "work +1 973-438-1486","CFO EMAIL": "Not found on Linkedin","CFO NUMBER": "Not found on Linkedin","IR Email": "","IR Contact": "","IR Page": ""},
  "ZOMD-CA": {"CEO": "Ido Almany","CFO": "Tsvika Adler","CEO EMAIL": "ido@zoomd.com","CEO NUMBER": "work 972 72-220-0555","CFO EMAIL": "tsvika@zoomd.com","CFO NUMBER": "","IR Email": "","IR Contact": "","IR Page": ""},
}

SCORE_FIELDS = ["CEO","CFO","CEO EMAIL","CEO NUMBER","CFO EMAIL","CFO NUMBER","IR Email","IR Contact","IR Page"]
TOTAL_FIELDS = len(GROUND_TRUTH) * len(SCORE_FIELDS)  # 837
PASS_THRESHOLD = 0.80

# Weighted scoring: CEO/CFO email+phone = 4x, names + IR = 1x
# Rationale: contact data (email/phone) is the primary deliverable.
# 4 contact fields × 4 pts + 5 other fields × 1 pt = 21 pts per ticker
FIELD_WEIGHTS: dict[str, int] = {
    "CEO EMAIL":   4,
    "CEO NUMBER":  4,
    "CFO EMAIL":   4,
    "CFO NUMBER":  4,
    "CEO":         1,
    "CFO":         1,
    "IR Email":    1,
    "IR Contact":  1,
    "IR Page":     1,
}
TOTAL_WEIGHTED = sum(FIELD_WEIGHTS.values()) * len(GROUND_TRUTH)  # 21 × 93 = 1953
PASS_THRESHOLD_WEIGHTED = int(TOTAL_WEIGHTED * PASS_THRESHOLD)    # 1562

# ─────────────────────────────────────────────────────────────────────────────
# ── Scoring logic ─────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────

_BLANK_ALIASES = {"", "not found", "n/a", "not on sql", "not on linkedin",
                   "not found on linkedin", "none"}

def _norm(v) -> str:
    return str(v or "").strip().lower()

def _digits_only(s: str) -> str:
    return re.sub(r"\D", "", s)

def _has_work_prefix(s: str) -> bool:
    return s.lower().startswith("work")

def score_field(actual, expected, field: str) -> tuple[bool, str]:
    """
    Return (pass, reason).
    """
    a = _norm(actual)
    e = _norm(expected)

    # Both blank → PASS
    if e in _BLANK_ALIASES and a in _BLANK_ALIASES:
        return True, "both blank"

    # Ground truth is blank/N/A, platform gave something → PASS (bonus)
    if e in _BLANK_ALIASES and a not in _BLANK_ALIASES:
        return True, "bonus find"

    # Ground truth has value, platform is blank → FAIL
    if e not in _BLANK_ALIASES and a in _BLANK_ALIASES:
        return False, "missing value"

    # ── Name fields (CEO / CFO) ────────────────────────────────────────────────
    if field in ("CEO", "CFO"):
        # Core name must match — ignore credentials (MBA, CPA, etc.)
        def _core(name: str) -> str:
            # Strip trailing credentials
            name = re.sub(
                r"\s+(jr\.?|sr\.?|ii|iii|iv|phd|ph\.d\.?|md|cpa|mba|cfa|jd|esq\.?)[\s,]*$",
                "", name, flags=re.IGNORECASE,
            ).strip()
            # Strip leading honorifics
            name = re.sub(r"^(mr\.?|ms\.?|mrs\.?|dr\.?|prof\.?)\s+", "", name, flags=re.IGNORECASE)
            # Normalize whitespace (yfinance sometimes returns double spaces)
            name = re.sub(r"\s+", " ", name).strip()
            return name.lower()
        return _core(a) == _core(e), "name match"

    # ── Email fields ───────────────────────────────────────────────────────────
    if "EMAIL" in field:
        # Exact match (case insensitive)
        if a == e:
            return True, "exact"
        # Platform found work email when ground truth is personal → PASS (upgrade)
        if "(no work provided)" in e and "(no work provided)" not in a and "@" in a:
            return True, "email upgrade to work"
        # Both "not on linkedin" / "not found" variants
        if any(kw in a for kw in ("not on", "not found")) and any(kw in e for kw in ("not on", "not found")):
            return True, "both not-found"
        return False, f"email mismatch: got '{actual}' expected '{expected}'"

    # ── Phone fields ───────────────────────────────────────────────────────────
    if "NUMBER" in field:
        # Not-found / Not-on-LinkedIn variants
        if any(kw in a for kw in ("not on", "not found")) and any(kw in e for kw in ("not on", "not found", "n/a")):
            return True, "both not-found"
        if a in _BLANK_ALIASES and e in {"n/a"}:
            return True, "both blank/n/a"

        # Digit comparison — must match
        a_digits = _digits_only(a)
        e_digits = _digits_only(e)
        if not a_digits or not e_digits:
            return a in _BLANK_ALIASES and e in _BLANK_ALIASES, "digit check"
        if a_digits != e_digits:
            return False, f"digits mismatch: got '{a_digits}' expected '{e_digits}'"

        # Work prefix check
        if _has_work_prefix(e) and not _has_work_prefix(a):
            return False, "missing 'work' prefix"
        if not _has_work_prefix(e) and _has_work_prefix(a):
            return False, "unexpected 'work' prefix"

        return True, "phone match"

    # ── IR fields ──────────────────────────────────────────────────────────────
    if field == "IR Email":
        if a == e:
            return True, "exact"
        if any(kw in a for kw in ("not on", "not found")) and any(kw in e for kw in ("not on", "not found")):
            return True, "both not-found"
        return False, f"ir email mismatch"

    if field == "IR Contact":
        if not e:
            return True, "both blank"
        # Core name match (firm name in parens is bonus, not required)
        e_core = re.sub(r"\s*\(.*\)\s*", "", e).strip().lower()
        a_core = re.sub(r"\s*\(.*\)\s*", "", a).strip().lower()
        if e_core and e_core in a_core:
            return True, "name match"
        # Word overlap ≥ 70%
        e_words = set(re.sub(r"[^a-z0-9]", " ", e_core).split())
        a_words = set(re.sub(r"[^a-z0-9]", " ", a_core).split())
        if e_words and len(a_words & e_words) / len(e_words) >= 0.7:
            return True, "partial name match"
        return False, f"ir contact mismatch"

    if field == "IR Page":
        def _norm_ir_url(u: str) -> str:
            u = (u or "").strip().lower().rstrip("/")
            u = re.sub(r"^https?://", "", u)
            u = re.sub(r"^www\.", "", u)
            return u

        def _base_domain(d: str) -> str:
            """'ir.backblaze.com' → 'backblaze.com'"""
            parts = d.split(".")
            return ".".join(parts[-2:]) if len(parts) >= 2 else d

        def _is_ir_url(u: str) -> bool:
            return any(kw in u for kw in ("invest", "/ir", "ir.", "shareholder", "relations"))

        a_n = _norm_ir_url(a)
        e_n = _norm_ir_url(e)
        if a_n == e_n:
            return True, "url match"

        a_domain = a_n.split("/")[0]
        e_domain = e_n.split("/")[0]

        # Exact subdomain match (e.g. "investors.1stdibs.com/contact" vs "investors.1stdibs.com/")
        if a_domain and e_domain and a_domain == e_domain:
            return True, "same ir domain"

        # Same base domain AND both URLs point to IR content
        # Handles: "www.backblaze.com/investor-relations" vs "ir.backblaze.com/"
        if _base_domain(a_domain) == _base_domain(e_domain) and _is_ir_url(a_n) and _is_ir_url(e_n):
            return True, "same base domain + ir content"

        return False, f"url mismatch: got '{a}' expected '{e}'"

    return a == e, "exact"


# ─────────────────────────────────────────────────────────────────────────────
# ── Pipeline runner ───────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────

SALESQL_DELAY = 1.5

def _clean_exec_name(raw: str) -> str:
    """Strip 'Mr./Ms./Dr.' prefix and normalize whitespace from yfinance names."""
    name = re.sub(r"^\s*(Mr\.?|Ms\.?|Mrs\.?|Miss\.?|Dr\.?|Prof\.?|Sir)\s+", "", raw, flags=re.IGNORECASE)
    name = re.sub(r"\s+", " ", name).strip()
    return name

def _role_label(title: str) -> str:
    t = title.lower()
    # Check for explicit CEO keywords anywhere in the title (handles "Co-Founder & CEO", "Chairman and CEO", etc.)
    if "chief executive" in t or "ceo" in re.split(r"[\s,&|/]+", t):
        return "CEO"
    if "chief financial" in t or "cfo" in re.split(r"[\s,&|/]+", t):
        return "CFO"
    if "chief operating" in t or "coo" in re.split(r"[\s,&|/]+", t):
        return "COO"
    if "president" in t and "vice" not in t:
        return "President"
    return title[:20]

def run_ticker(ticker: str, skip_linkedin: bool = False) -> dict:
    """Run full pipeline, return 21-col row dict."""
    try:
        company_info = resolve_ticker(ticker)
    except Exception as e:
        print(f"  [{ticker}] resolve error: {e}")
        return empty_row(ticker, "API Error")

    if not company_info:
        print(f"  [{ticker}] not found")
        return empty_row(ticker, "Not Found")

    company = company_info.get("company", "")
    website = company_info.get("website", "")   # used for domain-based SalesQL lookup
    targets  = company_info.get("targets", [])
    print(f"  [{ticker}] {company} — {len(targets)} execs")

    # Financial data
    try:
        financials = fetch_financials_safe(ticker)
    except Exception:
        financials = {}

    # Executives
    executives: list[dict] = []
    ceo_t = next((t for t in targets if _role_label(t["title"]) == "CEO"), None)
    cfo_t = next((t for t in targets if _role_label(t["title"]) == "CFO"), None)
    if not ceo_t and targets:
        ceo_t = targets[0]

    for target in [ceo_t, cfo_t]:
        if not target:
            continue
        name  = _clean_exec_name(target["name"])   # strip "Mr." prefix & double spaces
        title = target["title"]
        role  = _role_label(title)
        first, last = split_name(name)

        try:
            enrichment = salesql_empty("not_tried")
            li_url = None
            if not skip_linkedin:
                li_url = find_linkedin_url(name, company, title)
                print(f"    LinkedIn [{role}] {name}: {li_url or 'not found'}")
                time.sleep(SALESQL_DELAY)
                if li_url:
                    enrichment = enrich_by_url(li_url)
                    print(f"    SalesQL URL result: email={enrichment.get('best_email') or 'none'} phone={enrichment.get('phone') or 'none'} source={enrichment.get('source')}")
                    if enrichment.get("best_email") or enrichment.get("phone"):
                        enrichment["linkedin_url"] = li_url
                        executives.append({"role": role, "name": name, "title": title, "enrichment": enrichment})
                        continue
            time.sleep(SALESQL_DELAY)
            enrichment = search_by_name_with_variations(first, last, name, company, website=website)
            print(f"    SalesQL NAME [{role}] {name} @ {website or company}: email={enrichment.get('best_email') or 'none'} phone={enrichment.get('phone') or 'none'} phone_type={enrichment.get('phone_type') or 'none'} source={enrichment.get('source')}")
        except Exception as e:
            print(f"  [{ticker}] {role} enrichment error: {e}")
            enrichment = salesql_empty("error")

        executives.append({"role": role, "name": name, "title": title, "enrichment": enrichment})

    # ── Cross-domain retry ────────────────────────────────────────────────────
    # If any exec has a work email from a domain OTHER than the website domain
    # (e.g. CFO = anthonyg@mamamancinis.com, website = mamascreations.com),
    # retry any exec that has no work email using that brand domain.
    try:
        from lookup.salesql_enricher import _extract_domain as _sq_extract_domain
        from lookup.schema_builder import _is_personal_email as _sq_is_personal

        site_domain = _sq_extract_domain(website) if website else ""
        brand_domains: set[str] = set()
        for ex in executives:
            em = (ex["enrichment"].get("best_email") or "").strip()
            if em and "@" in em:
                d = em.split("@")[-1].lower()
                if (d != site_domain
                        and not _sq_is_personal(em)
                        and not d.endswith(".edu")
                        and len(d) > 4):
                    brand_domains.add(d)

        if brand_domains:
            for ex in executives:
                best = (ex["enrichment"].get("best_email") or "").strip()
                # Skip if already has a good work email
                if best and not _sq_is_personal(best) and not best.endswith(".edu"):
                    continue
                first, last = split_name(ex["name"])
                for bd in brand_domains:
                    time.sleep(SALESQL_DELAY)
                    retry = search_by_name_with_variations(
                        first, last, ex["name"], company, website=f"https://{bd}"
                    )
                    rm = (retry.get("best_email") or "").strip()
                    if rm and "@" in rm and not _sq_is_personal(rm) and not rm.endswith(".edu"):
                        ex["enrichment"] = retry
                        print(f"    [brand-domain retry] {ex['name']} → {rm} (via {bd})")
                        break
    except Exception as _cdr_err:
        pass

    # Email pattern inference
    try:
        _ep = [{"name": e["name"], "first_name": split_name(e["name"])[0],
                "last_name": split_name(e["name"])[1],
                "best_email": e["enrichment"].get("best_email","")} for e in executives]
        fill_missing_emails(_ep, website=company_info.get("website",""), verbose=False)
        for e, ep in zip(executives, _ep):
            if not e["enrichment"].get("best_email") and ep.get("best_email"):
                e["enrichment"]["best_email"] = ep["best_email"]
                e["enrichment"]["work_email"]  = ep["best_email"]
    except Exception:
        pass

    # IR data
    try:
        ir_data = find_ir_data(ticker=ticker, company=company, website=company_info.get("website",""))
        # IR SalesQL fallback
        if ir_data.get("ir_contact") and not ir_data.get("ir_email"):
            plain_ir = re.sub(r"\s*\(.*\)\s*$", "", ir_data["ir_contact"]).split(",")[0].strip()
            ir_first, ir_last = split_name(plain_ir)
            # Non-person word blacklist — last names that indicate a company/dept,
            # not a real person (e.g. "Design Firm", "IR Group", "Capital Partners")
            _NON_PERSON_LAST = {
                "firm", "group", "associates", "partners", "company", "corp",
                "inc", "llc", "ltd", "department", "dept", "team", "division",
                "office", "relations", "contact", "capital", "management",
                "services", "solutions", "communications", "international",
            }
            # Require first+last, and last name must not be a non-person word
            if ir_first and ir_last and ir_last.lower().rstrip(".,") not in _NON_PERSON_LAST:
                try:
                    ir_en = search_by_name_with_variations(ir_first, ir_last, plain_ir, company, website=website)
                    em = ir_en.get("best_email") or ir_en.get("work_email") or ""
                    if em:
                        ir_data["ir_email"] = em
                except Exception:
                    pass
    except Exception:
        ir_data = {}

    # Exchange
    try:
        import yfinance as yf
        company_info["exchange"] = (yf.Ticker(ticker).info or {}).get("exchange","")
    except Exception:
        pass

    return build_row(ticker=ticker, company_info=company_info,
                     financials=financials, executives=executives, ir_data=ir_data)


# ─────────────────────────────────────────────────────────────────────────────
# ── Scoring runner ────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────

def score_row(row: dict) -> list[dict]:
    """Score one output row against ground truth. Returns list of field results."""
    ticker = row.get("Ticker","")
    gt     = GROUND_TRUTH.get(ticker)
    if not gt:
        return []

    results = []
    for field in SCORE_FIELDS:
        actual   = row.get(field, "")
        expected = gt.get(field, "")
        passed, reason = score_field(actual, expected, field)
        results.append({
            "Ticker":   ticker,
            "Field":    field,
            "Pass":     passed,
            "Weight":   FIELD_WEIGHTS.get(field, 1),
            "Actual":   str(actual or ""),
            "Expected": str(expected or ""),
            "Reason":   reason,
        })
    return results


# ─────────────────────────────────────────────────────────────────────────────
# ── xlsx report writer ────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────

def write_report(all_rows: list[dict], all_scores: list[dict], out_dir: Path) -> None:
    # ── Full scored output ────────────────────────────────────────────────────
    df_rows   = pd.DataFrame(all_rows, columns=COLUMN_ORDER)
    df_scores = pd.DataFrame(all_scores)

    # ── Mismatch-only ─────────────────────────────────────────────────────────
    df_mis = df_scores[df_scores["Pass"] == False].copy()

    # ── Per-ticker accuracy ────────────────────────────────────────────────────
    total = len(df_scores)
    passed = df_scores["Pass"].sum()
    pct    = round(100 * passed / total, 1) if total else 0.0

    ticker_acc = (
        df_scores.groupby("Ticker")
        .agg(Passed=("Pass","sum"), Total=("Pass","count"))
        .assign(Accuracy=lambda d: (d["Passed"]/d["Total"]*100).round(1))
        .reset_index()
        .sort_values("Accuracy")
    )

    # Write xlsx
    out_xlsx = out_dir / "validate_results.xlsx"
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as wr:
        df_rows.to_excel(wr, sheet_name="Output", index=False)
        df_scores.to_excel(wr, sheet_name="Scored Fields", index=False)
        df_mis.to_excel(wr, sheet_name="Mismatches", index=False)
        ticker_acc.to_excel(wr, sheet_name="Per-Ticker Accuracy", index=False)

        # Colour-code Scored Fields
        ws = wr.sheets["Scored Fields"]
        green = PatternFill("solid", fgColor="0D2818")
        red   = PatternFill("solid", fgColor="2D1111")
        for row in ws.iter_rows(min_row=2):
            pass_val = str(row[2].value)
            fill = green if pass_val == "True" else red
            for cell in row:
                cell.fill = fill

    print(f"\n  Saved: {out_xlsx}")

    mis_xlsx = out_dir / "validate_mismatches.xlsx"
    df_mis.to_excel(mis_xlsx, index=False)
    print(f"  Saved: {mis_xlsx}")

    return pct, passed, total


# ─────────────────────────────────────────────────────────────────────────────
# ── Main ──────────────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="93-ticker accuracy validation")
    parser.add_argument("tickers", nargs="*", help="Specific tickers to test (default: all 93)")
    parser.add_argument("--skip-linkedin", action="store_true",
                        help="Skip LinkedIn URL lookup (faster, name-only SalesQL)")
    parser.add_argument("--delay", type=float, default=1.5,
                        help="SalesQL delay in seconds (default: 1.5)")
    parser.add_argument("--resume", action="store_true",
                        help="Skip tickers already in validate_progress.json")
    args = parser.parse_args()

    global SALESQL_DELAY
    SALESQL_DELAY = args.delay

    tickers_to_run = [t.upper() for t in args.tickers] if args.tickers else list(GROUND_TRUTH.keys())

    # Resume support
    progress_file = APP_DIR / "validate_progress.json"
    done_rows: list[dict] = []
    if args.resume and progress_file.exists():
        with open(progress_file) as f:
            done_rows = json.load(f)
        done_tickers = {r["Ticker"] for r in done_rows}
        tickers_to_run = [t for t in tickers_to_run if t not in done_tickers]
        print(f"  Resuming: {len(done_rows)} already done, {len(tickers_to_run)} remaining")

    print(f"\n{'='*60}")
    print(f"  IB Research Engine — Accuracy Validation")
    print(f"  {len(tickers_to_run)} tickers to process")
    print(f"  Pass threshold: {PASS_THRESHOLD*100:.0f}% weighted  ({PASS_THRESHOLD_WEIGHTED}/{TOTAL_WEIGHTED} pts)")
    print(f"{'='*60}\n")

    all_rows  = list(done_rows)
    n = len(tickers_to_run)

    for i, ticker in enumerate(tickers_to_run, 1):
        print(f"[{i}/{n}] {ticker}")
        try:
            row = run_ticker(ticker, skip_linkedin=args.skip_linkedin)
        except Exception as e:
            print(f"  ERROR: {e}")
            row = empty_row(ticker, "API Error")

        all_rows.append(row)

        # Save progress incrementally
        with open(progress_file, "w") as f:
            json.dump(all_rows, f, default=str)

        if i < n:
            time.sleep(0.3)

    # Score
    all_scores: list[dict] = []
    for row in all_rows:
        all_scores.extend(score_row(row))

    # Report — unweighted + weighted
    total  = len(all_scores)
    passed = sum(1 for s in all_scores if s["Pass"])
    pct    = round(100 * passed / total, 1) if total else 0.0

    total_w  = sum(s.get("Weight", 1) for s in all_scores)
    passed_w = sum(s.get("Weight", 1) for s in all_scores if s["Pass"])
    pct_w    = round(100 * passed_w / total_w, 1) if total_w else 0.0

    print(f"\n{'='*60}")
    print(f"  RESULTS (unweighted):  {pct}%  ({passed}/{total} fields)")
    print(f"  RESULTS (weighted):    {pct_w}%  ({passed_w}/{total_w} pts)")
    print(f"    — CEO/CFO email+phone = 4pts each; names/IR = 1pt each")
    symbol = "✅ PASS" if pct_w >= PASS_THRESHOLD * 100 else "❌ FAIL"
    print(f"  {symbol}  (weighted threshold: {PASS_THRESHOLD*100:.0f}%)")
    print(f"{'='*60}")

    # Per-ticker summary
    from collections import defaultdict
    ticker_totals: dict = defaultdict(lambda: {"pass": 0, "total": 0})
    for s in all_scores:
        t = s["Ticker"]
        ticker_totals[t]["total"] += 1
        if s["Pass"]:
            ticker_totals[t]["pass"] += 1

    print(f"\n  Per-ticker (lowest first):")
    for t, v in sorted(ticker_totals.items(), key=lambda x: x[1]["pass"]/x[1]["total"]):
        pct_t = round(100 * v["pass"] / v["total"], 0)
        flag  = "✅" if pct_t >= 80 else "⚠" if pct_t >= 50 else "❌"
        print(f"    {flag} {t:12s} {pct_t:3.0f}%  ({v['pass']}/{v['total']})")

    # Mismatch summary
    mismatches = [s for s in all_scores if not s["Pass"]]
    print(f"\n  Mismatches by field ({len(mismatches)} total):")
    from collections import Counter
    field_fails = Counter(s["Field"] for s in mismatches)
    for field, cnt in sorted(field_fails.items(), key=lambda x: -x[1]):
        print(f"    {field:20s}: {cnt} fails")

    # Write xlsx reports
    write_report(all_rows, all_scores, APP_DIR)

    # Clean up progress file on successful completion
    if progress_file.exists():
        try:
            progress_file.unlink()
        except (PermissionError, OSError):
            # On some filesystems (mounted volumes) unlink is not permitted;
            # overwrite with empty list so --resume won't re-use stale data.
            try:
                progress_file.write_text("[]")
            except Exception:
                pass

    return 0 if pct >= PASS_THRESHOLD * 100 else 1


if __name__ == "__main__":
    sys.exit(main())
