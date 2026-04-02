"""
excel_writer.py
Build a clean, formatted Excel workbook from enriched_leads.csv.

Two sheets:
  1. Executive Contacts  — Name, Role, LinkedIn, Email, Phone (clean view)
  2. Full Data           — all raw fields for reference
"""
from __future__ import annotations
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colour palette ────────────────────────────────────────────────────────────
NAVY      = "1F3864"
NAVY_LITE = "2E4A7A"
GRAY_HDR  = "404040"
GRAY_ROW  = "F2F2F2"
WHITE     = "FFFFFF"
GREEN     = "1E8449"
ORANGE    = "E67E22"
LINK_BLUE = "1155CC"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _border() -> Border:
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr_font(color: str = "FFFFFF") -> Font:
    return Font(name="Arial", bold=True, color=color, size=10)


def _cell_font(bold: bool = False, color: str = "000000", size: int = 9) -> Font:
    return Font(name="Arial", bold=bold, color=color, size=size)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center")


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=False)


# ── Name cleaner ──────────────────────────────────────────────────────────────
_PREFIX_RE = re.compile(r'^(Mr\.|Ms\.|Mrs\.|Dr\.|Prof\.)\s*', re.IGNORECASE)
_SUFFIX_RE = re.compile(r'\s+(M\.D\.|CPA.*|J\.D\.|MBA.*|Ph\.D\.).*$', re.IGNORECASE)


def _clean_name(name: str) -> str:
    name = _PREFIX_RE.sub("", name.strip())
    name = _SUFFIX_RE.sub("", name)
    return name.strip()


# ── Main builder ──────────────────────────────────────────────────────────────

def build_excel(csv_path: str, xlsx_path: str) -> dict:
    """
    Read *csv_path* and write a formatted Excel workbook to *xlsx_path*.
    Returns a summary dict with record counts.
    """
    df = pd.read_csv(csv_path, dtype=str).fillna("")
    df["_clean_name"] = df["executive_name"].apply(_clean_name)

    # ── Sheet 1 data ──────────────────────────────────────────────────────────
    clean_df = pd.DataFrame({
        "Ticker":       df["ticker"],
        "Company":      df["company"],
        "Role":         df["role"],
        "Name":         df["_clean_name"],
        "Title":        df["title"],
        "LinkedIn URL": df["linkedin_url"],
        "Email":        df["best_email"],
        "Phone":        df["phone"],
        "Industry":     df["industry"],
        "Location":     df.apply(
            lambda r: ", ".join(filter(None, [r["city"], r["state"], r["country"]])), axis=1
        ),
    })

    wb = Workbook()

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1 — Executive Contacts (clean view)
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Executive Contacts"

    # Title
    ws1.merge_cells("A1:J1")
    c = ws1["A1"]
    c.value     = "IB Executive Contact Database"
    c.font      = Font(name="Arial", bold=True, size=14, color=WHITE)
    c.fill      = _fill(NAVY)
    c.alignment = _center()
    ws1.row_dimensions[1].height = 28

    # Subtitle
    ws1.merge_cells("A2:J2")
    c = ws1["A2"]
    c.value     = (f"CEO / CFO / President contacts — "
                   f"{len(clean_df)} records across {clean_df['Ticker'].nunique()} companies")
    c.font      = Font(name="Arial", italic=True, size=9, color=WHITE)
    c.fill      = _fill(NAVY_LITE)
    c.alignment = _center()
    ws1.row_dimensions[2].height = 16

    ws1.row_dimensions[3].height = 5  # spacer

    # Headers (row 4)
    HEADERS = ["Ticker", "Company", "Role", "Name", "Title",
               "LinkedIn URL", "Email", "Phone", "Industry", "Location"]
    for ci, hdr in enumerate(HEADERS, 1):
        cell = ws1.cell(row=4, column=ci, value=hdr)
        cell.font      = _hdr_font()
        cell.fill      = _fill(NAVY)
        cell.alignment = _center()
        cell.border    = _border()
    ws1.row_dimensions[4].height = 18

    # Data rows
    for ri, (_, row) in enumerate(clean_df.iterrows(), 5):
        bg = GRAY_ROW if ri % 2 == 0 else WHITE
        vals = [row["Ticker"], row["Company"], row["Role"], row["Name"],
                row["Title"], row["LinkedIn URL"], row["Email"], row["Phone"],
                row["Industry"], row["Location"]]

        for ci, val in enumerate(vals, 1):
            cell = ws1.cell(row=ri, column=ci, value=val or "")
            cell.fill   = _fill(bg)
            cell.border = _border()
            cell.alignment = _left()

            if ci == 3:  # Role — colour by seniority
                role_up = str(val).upper()
                if role_up == "CEO":
                    cell.font = Font(name="Arial", bold=True, size=9, color=NAVY)
                elif role_up == "CFO":
                    cell.font = Font(name="Arial", bold=True, size=9, color=GREEN)
                elif role_up in ("COO", "PRESIDENT"):
                    cell.font = Font(name="Arial", bold=True, size=9, color=ORANGE)
                else:
                    cell.font = _cell_font()

            elif ci == 6 and val:  # LinkedIn URL → hyperlink
                cell.font      = Font(name="Arial", size=9, color=LINK_BLUE, underline="single")
                cell.hyperlink = val

            elif ci == 7 and val:  # Email → green
                cell.font = Font(name="Arial", size=9, color=GREEN)

            elif not val:  # Empty → dim dash
                cell.font  = Font(name="Arial", size=9, color="BBBBBB")
                cell.value = "—"

            else:
                cell.font = _cell_font()

        ws1.row_dimensions[ri].height = 15

    # Column widths
    for ci, w in enumerate([8, 26, 10, 22, 34, 44, 32, 18, 22, 22], 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    ws1.freeze_panes = "A5"
    ws1.auto_filter.ref = f"A4:{get_column_letter(len(HEADERS))}{len(clean_df) + 4}"

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2 — Full Data (all raw CSV fields)
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Full Data")
    raw_cols = [c for c in df.columns if c != "_clean_name"]

    for ci, hdr in enumerate(raw_cols, 1):
        cell = ws2.cell(row=1, column=ci, value=hdr.replace("_", " ").title())
        cell.font      = _hdr_font()
        cell.fill      = _fill(GRAY_HDR)
        cell.alignment = _center()
        cell.border    = _border()
    ws2.row_dimensions[1].height = 16

    for ri, (_, row) in enumerate(df[raw_cols].iterrows(), 2):
        bg = GRAY_ROW if ri % 2 == 0 else WHITE
        for ci, col in enumerate(raw_cols, 1):
            cell = ws2.cell(row=ri, column=ci, value=row[col] or "")
            cell.font      = _cell_font()
            cell.fill      = _fill(bg)
            cell.alignment = _left()
            cell.border    = _border()
        ws2.row_dimensions[ri].height = 14

    for ci, col in enumerate(raw_cols, 1):
        max_len = max(len(col), df[col].astype(str).str.len().max() if len(df) else 10)
        ws2.column_dimensions[get_column_letter(ci)].width = min(max_len + 3, 42)

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(raw_cols))}{len(df) + 1}"

    wb.save(xlsx_path)

    return {
        "records":    len(clean_df),
        "tickers":    int(clean_df["Ticker"].nunique()),
        "with_email": int((clean_df["Email"] != "").sum()),
        "with_phone": int((clean_df["Phone"] != "").sum()),
        "path":       xlsx_path,
    }
