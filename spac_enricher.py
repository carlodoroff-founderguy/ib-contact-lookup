#!/usr/bin/env python3
"""
spac_enricher.py  —  Standalone SPAC contact enrichment runner

Reads spac_data/spac_roster.json, enriches CEO/CFO contacts via
SalesQL + LinkedIn, saves the updated roster, and exports an Excel report.

Performance:
  - 5 parallel workers via ThreadPoolExecutor (~5x speedup)
  - Smarter delays: 0.8s SalesQL, 1.2s DuckDuckGo
  - Person cache: same exec across multiple SPACs → one API call
  - Incremental save after each ticker (crash-safe, --resume friendly)
  - Auto-sorted by urgency (most urgent SPACs finish first)

Usage:
  python spac_enricher.py                       # all SPACs that need enrichment
  python spac_enricher.py LOKV CCII RANG        # specific tickers only
  python spac_enricher.py --urgent              # only SPACs with <90 days remaining
  python spac_enricher.py --resume              # skip already-completed tickers
  python spac_enricher.py --urgent --resume     # combine flags
  python spac_enricher.py --dry-run             # preview what would run (no API calls)
  python spac_enricher.py --workers 3           # use 3 parallel workers (default: 5)
"""

from __future__ import annotations

import argparse
import sys
import time
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

from lookup.spac_contact_lookup import (
    load_roster,
    save_roster,
    needs_enrichment,
    get_urgency,
    enrich_spac_ticker,
    reset_cache,
)


# ── Constants ────────────────────────────────────────────────────────────────

_ROSTER_DIR = Path(__file__).resolve().parent / "spac_data"
_ROSTER_PATH = _ROSTER_DIR / "spac_roster.json"

MAX_WORKERS = 5   # default parallel workers (do not exceed 8)

# ── Thread-safe counters ─────────────────────────────────────────────────────

_counter_lock = threading.Lock()
_enriched = 0
_failed = 0
_no_result = 0
_processed = 0


def _inc(counter_name: str):
    global _enriched, _failed, _no_result, _processed
    with _counter_lock:
        if counter_name == "enriched":
            _enriched += 1
        elif counter_name == "failed":
            _failed += 1
        elif counter_name == "no_result":
            _no_result += 1
        _processed += 1


# ── Excel export ─────────────────────────────────────────────────────────────

# Urgency colour fills
_FILL_RED    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") if _HAS_OPENPYXL else None
_FILL_ORANGE = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid") if _HAS_OPENPYXL else None
_FILL_YELLOW = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid") if _HAS_OPENPYXL else None
_FILL_GREEN  = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid") if _HAS_OPENPYXL else None

_HEADER_FILL = PatternFill(start_color="1C2B3A", end_color="1C2B3A", fill_type="solid") if _HAS_OPENPYXL else None
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11) if _HAS_OPENPYXL else None

_THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
) if _HAS_OPENPYXL else None

_EXCEL_COLUMNS = [
    "Ticker",
    "Company",
    "Deadline",
    "Days Remaining",
    "Urgency",
    "CEO Name",
    "CEO Email",
    "CEO Phone",
    "CFO Name",
    "CFO Email",
    "CFO Phone",
    "Source",
    "Notes",
]


def _urgency_fill(days):
    """Return the openpyxl fill based on days remaining."""
    if not _HAS_OPENPYXL or days is None:
        return None
    if days < 30:
        return _FILL_RED
    if days <= 90:
        return _FILL_ORANGE
    if days <= 180:
        return _FILL_YELLOW
    return None


def export_excel(roster: dict, output_path: Path) -> Path:
    """Export roster to colour-coded Excel report."""
    if not _HAS_OPENPYXL:
        print("  ⚠ openpyxl not installed — skipping Excel export")
        return output_path

    wb = openpyxl.Workbook()

    # ── Sheet 1: All SPACs ───────────────────────────────────────────────────
    ws = wb.active
    ws.title = "SPAC Roster"

    for col_idx, header in enumerate(_EXCEL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _THIN_BORDER

    sorted_tickers = sorted(
        roster.keys(),
        key=lambda t: roster[t].get("days_remaining") or 9999,
    )

    for row_idx, ticker in enumerate(sorted_tickers, 2):
        d = roster[ticker]
        days = d.get("days_remaining")
        urgency = get_urgency(d)
        row_fill = _urgency_fill(days)

        values = [
            ticker,
            d.get("company", ""),
            d.get("deadline", ""),
            days,
            urgency,
            d.get("ceo_name", ""),
            d.get("ceo_email", ""),
            d.get("ceo_phone", ""),
            d.get("cfo_name", ""),
            d.get("cfo_email", ""),
            d.get("cfo_phone", ""),
            d.get("source", ""),
            d.get("notes", ""),
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            if row_fill:
                cell.fill = row_fill
            if col_idx in (7, 10) and val:
                cell.font = Font(name="Calibri", bold=True, size=11)

    _widths = [10, 35, 14, 14, 12, 28, 30, 18, 28, 30, 18, 20, 35]
    for i, w in enumerate(_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # ── Sheet 2: Missing contacts only ───────────────────────────────────────
    ws2 = wb.create_sheet("Missing Contacts")
    for col_idx, header in enumerate(_EXCEL_COLUMNS, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _THIN_BORDER

    row_idx2 = 2
    for ticker in sorted_tickers:
        d = roster[ticker]
        if not needs_enrichment(d):
            continue
        days = d.get("days_remaining")
        urgency = get_urgency(d)
        row_fill = _urgency_fill(days)
        values = [
            ticker,
            d.get("company", ""),
            d.get("deadline", ""),
            days,
            urgency,
            d.get("ceo_name", ""),
            d.get("ceo_email", ""),
            d.get("ceo_phone", ""),
            d.get("cfo_name", ""),
            d.get("cfo_email", ""),
            d.get("cfo_phone", ""),
            d.get("source", ""),
            d.get("notes", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws2.cell(row=row_idx2, column=col_idx, value=val)
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            if row_fill:
                cell.fill = row_fill
        row_idx2 += 1

    for i, w in enumerate(_widths, 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    wb.save(str(output_path))
    return output_path


# ── CLI ──────────────────────────────────────────────────────────────────────

_print_lock = threading.Lock()


def _log(msg: str, level: str = "dim"):
    """Thread-safe colour-coded terminal logger."""
    colours = {
        "ok":   "\033[92m",
        "warn": "\033[93m",
        "err":  "\033[91m",
        "info": "\033[96m",
        "dim":  "\033[90m",
    }
    reset = "\033[0m"
    prefix = colours.get(level, "")
    with _print_lock:
        print(f"{prefix}{msg}{reset}")


def _enrich_one(ticker: str, roster: dict, args) -> str:
    """
    Enrich a single ticker — called from ThreadPoolExecutor.
    Updates roster[ticker] in-place, saves incrementally, returns ticker.
    """
    try:
        roster[ticker] = enrich_spac_ticker(
            ticker=ticker,
            spac_data=roster[ticker],
            delay=args.delay,
            skip_linkedin=args.no_linkedin,
            log_fn=_log,
        )

        d = roster[ticker]
        has_ceo = bool(d.get("ceo_email"))
        has_cfo = bool(d.get("cfo_email"))
        if has_ceo or has_cfo:
            _inc("enriched")
        else:
            _inc("no_result")

    except Exception as e:
        _log(f"  ✗ Error enriching {ticker}: {e}", "err")
        _inc("failed")

    # FIX 6: Save roster after each ticker (thread-safe, crash-safe)
    save_roster(roster, _ROSTER_PATH)

    with _counter_lock:
        current = _processed
    total = len([t for t in roster if True])  # just need any count
    _log(f"  [{current}/{args._total}] {ticker} done", "info")

    return ticker


def main():
    global _enriched, _failed, _no_result, _processed

    parser = argparse.ArgumentParser(
        description="SPAC Contact Enrichment — parallel batch enrichment of CEO/CFO contacts",
    )
    parser.add_argument(
        "tickers", nargs="*", type=str,
        help="Specific tickers to enrich (default: all that need enrichment)",
    )
    parser.add_argument(
        "--urgent", action="store_true",
        help="Only SPACs with <90 days remaining",
    )
    parser.add_argument(
        "--resume", action="store_true",
        help="Skip tickers that already have both CEO and CFO email",
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Preview which tickers would run (no API calls)",
    )
    parser.add_argument(
        "--delay", type=float, default=0.8,
        help="Seconds between SalesQL API calls (default: 0.8)",
    )
    parser.add_argument(
        "--workers", type=int, default=MAX_WORKERS,
        help=f"Parallel workers (default: {MAX_WORKERS}, max: 8)",
    )
    parser.add_argument(
        "--no-linkedin", action="store_true",
        help="Skip LinkedIn search steps (SalesQL name search only)",
    )
    parser.add_argument(
        "--no-excel", action="store_true",
        help="Skip Excel export after enrichment",
    )

    args = parser.parse_args()

    # Clamp workers to safe range
    args.workers = max(1, min(8, args.workers))

    # ── Load roster ──────────────────────────────────────────────────────────
    roster = load_roster(_ROSTER_PATH)
    if not roster:
        print("✗ No roster found at", _ROSTER_PATH)
        sys.exit(1)

    _log(f"Loaded {len(roster)} SPACs from roster", "info")

    # ── Filter tickers ───────────────────────────────────────────────────────
    if args.tickers:
        target_tickers = [t.strip().upper() for t in args.tickers]
        missing = [t for t in target_tickers if t not in roster]
        if missing:
            _log(f"⚠ Tickers not in roster: {', '.join(missing)}", "warn")
        target_tickers = [t for t in target_tickers if t in roster]
    else:
        target_tickers = list(roster.keys())

    # --urgent filter
    if args.urgent:
        target_tickers = [
            t for t in target_tickers
            if (roster[t].get("days_remaining") or 9999) < 90
        ]
        _log(f"Urgent filter: {len(target_tickers)} SPACs with <90 days", "info")

    # --resume filter
    if args.resume:
        before = len(target_tickers)
        target_tickers = [
            t for t in target_tickers
            if needs_enrichment(roster[t])
        ]
        skipped = before - len(target_tickers)
        if skipped:
            _log(f"Resume filter: skipped {skipped} already-complete tickers", "info")

    if not target_tickers:
        _log("No tickers to process after filtering.", "warn")
        sys.exit(0)

    # ── FIX 5: Sort by urgency (most urgent first — always) ─────────────────
    target_tickers.sort(key=lambda t: roster[t].get("days_remaining") or 9999)

    # ── Dry run ──────────────────────────────────────────────────────────────
    if args.dry_run:
        _log(f"\n{'='*60}", "info")
        _log(f"DRY RUN — {len(target_tickers)} tickers would be enriched:", "info")
        _log(f"Workers: {args.workers} parallel", "info")
        _log(f"{'='*60}\n", "info")
        for t in target_tickers:
            d = roster[t]
            urgency = get_urgency(d)
            needs = []
            if d.get("ceo_name") and not d.get("ceo_email"):
                needs.append("CEO email")
            if d.get("cfo_name") and not d.get("cfo_email"):
                needs.append("CFO email")
            needs_str = ", ".join(needs) if needs else "(already complete)"
            _log(f"  {t:8s}  {d.get('company',''):35s}  [{urgency:10s}]  needs: {needs_str}", "dim")
        _log(f"\n{'='*60}", "info")
        sys.exit(0)

    # ── FIX 4: Reset person cache for this run ───────────────────────────────
    reset_cache()

    # ── Enrichment — parallel with ThreadPoolExecutor ────────────────────────
    total = len(target_tickers)
    args._total = total  # stash for progress logging

    _log(f"\n{'='*60}", "info")
    _log(f"Starting enrichment of {total} SPACs", "info")
    _log(f"Workers: {args.workers} parallel", "info")
    _log(f"SalesQL delay: {args.delay}s", "info")
    _log(f"LinkedIn: {'disabled' if args.no_linkedin else 'enabled'}", "info")
    _log(f"{'='*60}\n", "info")

    _enriched = 0
    _failed = 0
    _no_result = 0
    _processed = 0

    t_start = time.time()

    with ThreadPoolExecutor(max_workers=args.workers) as executor:
        futures = {
            executor.submit(_enrich_one, ticker, roster, args): ticker
            for ticker in target_tickers
        }
        for future in as_completed(futures):
            ticker = futures[future]
            try:
                future.result()  # propagate exceptions
            except Exception as e:
                _log(f"  ✗ Unexpected error for {ticker}: {e}", "err")

    elapsed = time.time() - t_start

    # ── Summary ──────────────────────────────────────────────────────────────
    _log(f"\n{'='*60}", "info")
    _log(f"Enrichment complete in {elapsed:.1f}s", "info")
    _log(f"  Processed: {total}", "info")
    _log(f"  With contacts: {_enriched}", "ok")
    _log(f"  No contacts found: {_no_result}", "warn")
    _log(f"  Errors: {_failed}", "err" if _failed else "dim")
    _log(f"{'='*60}\n", "info")

    # ── Count overall stats ──────────────────────────────────────────────────
    total_spacs = len(roster)
    total_with_ceo = sum(1 for d in roster.values() if d.get("ceo_email"))
    total_with_cfo = sum(1 for d in roster.values() if d.get("cfo_email"))
    total_complete = sum(
        1 for d in roster.values()
        if (not d.get("ceo_name") or d.get("ceo_email"))
        and (not d.get("cfo_name") or d.get("cfo_email"))
    )
    _log(f"Overall roster stats:", "info")
    _log(f"  Total SPACs: {total_spacs}", "dim")
    _log(f"  CEO emails found: {total_with_ceo}/{total_spacs}", "dim")
    _log(f"  CFO emails found: {total_with_cfo}/{total_spacs}", "dim")
    _log(f"  Fully complete: {total_complete}/{total_spacs}", "dim")

    # ── Excel export ─────────────────────────────────────────────────────────
    if not args.no_excel:
        date_str = datetime.now().strftime("%Y-%m-%d")
        xlsx_path = _ROSTER_DIR / f"spac_results_{date_str}.xlsx"
        try:
            export_excel(roster, xlsx_path)
            _log(f"\n✓ Excel report saved: {xlsx_path}", "ok")
        except Exception as e:
            _log(f"\n✗ Excel export failed: {e}", "err")

    _log("Done.", "info")


if __name__ == "__main__":
    main()
