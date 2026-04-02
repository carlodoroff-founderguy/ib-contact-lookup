"""
app.py  —  Curvature Securities IB Intelligence Platform
Input tickers → company financials + CEO/CFO/IR contacts → export to Excel

Run:  streamlit run app.py

CLEAN REWRITE — single unified CSS, no duplicate widgets, consistent state keys.
"""
from __future__ import annotations

import io
import os
import re
import time
import threading
import concurrent.futures as _cf
from datetime import datetime, date
from pathlib import Path

import streamlit as st
import pandas as pd

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# ── Lookup pipeline ────────────────────────────────────────────────────────────
from lookup.ticker_resolver   import resolve_ticker, split_name
from lookup.linkedin_finder   import find_linkedin_url
from lookup.salesql_enricher  import (
    search_by_name_with_variations,
    search_by_name_and_company,
    enrich_by_url,
    _empty as salesql_empty,
)
from lookup.financial_fetcher import fetch_financials_safe
from lookup.ir_finder         import find_ir_data
from lookup.schema_builder    import build_row, empty_row, COLUMN_ORDER
from lookup.email_pattern     import fill_missing_emails
from lookup.bouncer_verifier  import verify_row_emails
from lookup.spac_detector     import detect_spac, resolve_spac_domain, get_edgar_filing_url
from lookup.spac_contact_lookup import (
    load_roster, save_roster, needs_enrichment, get_urgency,
    enrich_spac_ticker, reset_cache,
)


# ══════════════════════════════════════════════════════════════════════════════
# ── Page config (MUST be first Streamlit call) ────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="Curvature Securities — IB Intelligence",
    page_icon="◈",
    layout="wide",
    initial_sidebar_state="expanded",
)


# ══════════════════════════════════════════════════════════════════════════════
# ── CSS — single unified block ────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════

st.markdown("""
<style>
/* ── Fonts ─────────────────────────────────────────────────────────────────── */
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:ital,opsz,wght@0,9..40,300;0,9..40,400;0,9..40,500;0,9..40,600;0,9..40,700;1,9..40,400&family=Playfair+Display:wght@400;600;700&family=JetBrains+Mono:wght@400;500&display=swap');

/* ── Strip Streamlit chrome ────────────────────────────────────────────────── */
#MainMenu { visibility:hidden !important; }
footer { display:none !important; }
[data-testid="stToolbar"], .stDeployButton { display:none !important; }

/* ── Header — transparent so sidebar toggle stays in DOM ───────────────────── */
header[data-testid="stHeader"] {
    background-color:transparent !important;
    box-shadow:none !important;
    backdrop-filter:none !important;
    -webkit-backdrop-filter:none !important;
    border:none !important;
}

/* ── Sidebar toggle — always visible, navy tab on left edge ────────────────── */
button[data-testid="collapsedControl"],
[data-testid="collapsedControl"] button {
    display:flex !important;
    visibility:visible !important;
    opacity:1 !important;
    background-color:#1C2B3A !important;
    color:#FFFFFF !important;
    border-radius:0 8px 8px 0 !important;
    top:50% !important;
    position:fixed !important;
    left:0 !important;
    z-index:999999 !important;
    width:24px !important;
    height:48px !important;
    border:none !important;
    cursor:pointer !important;
    pointer-events:auto !important;
}
button[data-testid="collapsedControl"]:hover,
[data-testid="collapsedControl"] button:hover {
    background-color:#B8960C !important;
}
/* Sidebar close button inside the sidebar */
button[data-testid="stSidebarCollapseButton"],
[data-testid="stSidebarCollapseButton"] button {
    display:flex !important;
    visibility:visible !important;
    opacity:1 !important;
}

/* ── Gold rule at very top ─────────────────────────────────────────────────── */
body::before {
    content:''; display:block; position:fixed;
    top:0; left:0; right:0; height:3px;
    background:#B8960C; z-index:99999;
}

/* ── Page base ─────────────────────────────────────────────────────────────── */
html, body, .stApp {
    background:#FAFAF8 !important;
    color:#1C2B3A !important;
    font-family:'DM Sans',sans-serif !important;
}
.block-container {
    padding:2.5rem 2.5rem 5rem !important;
    max-width:1800px !important;
}

/* ── Sidebar ───────────────────────────────────────────────────────────────── */
section[data-testid="stSidebar"] {
    background:#FFFFFF !important;
    border-right:1px solid #E8E4DC !important;
}
section[data-testid="stSidebar"] * { color:#1C2B3A !important; }
section[data-testid="stSidebar"] p,
section[data-testid="stSidebar"] small { color:#6B7280 !important; }

/* ── Text inputs & textareas ───────────────────────────────────────────────── */
.stTextInput input, .stTextArea textarea {
    background:#FFFFFF !important;
    border:1px solid #E8E4DC !important;
    border-radius:8px !important;
    color:#1C2B3A !important;
    font-family:'JetBrains Mono',monospace !important;
    font-size:13px !important;
}
.stTextArea textarea { min-height:120px !important; }
.stTextInput input:focus, .stTextArea textarea:focus {
    border-color:#B8960C !important;
    box-shadow:0 0 0 3px rgba(184,150,12,0.15) !important;
    outline:none !important;
}
.stTextInput label, .stTextArea label {
    color:#9CA3AF !important;
    font-size:11px !important;
    font-family:'DM Sans',sans-serif !important;
    text-transform:uppercase !important;
    letter-spacing:0.12em !important;
    font-weight:500 !important;
}
.stTextInput input::placeholder, .stTextArea textarea::placeholder {
    color:#9CA3AF !important;
}

/* ── PRIMARY button — Run Research (navy, gold left accent) ────────────────── */
.stButton button[kind="primary"] {
    background:#1C2B3A !important;
    color:#FFFFFF !important;
    border:none !important;
    border-left:4px solid #B8960C !important;
    font-family:'DM Sans',sans-serif !important;
    font-weight:700 !important;
    font-size:13px !important;
    letter-spacing:0.10em !important;
    text-transform:uppercase !important;
    height:52px !important;
    border-radius:8px !important;
    transition:background 0.18s ease, box-shadow 0.18s ease !important;
}
.stButton button[kind="primary"]:hover {
    background:#2D4560 !important;
    box-shadow:0 4px 16px rgba(28,43,58,0.20) !important;
}
.stButton button[kind="primary"]:disabled {
    background:#1C2B3A !important;
    color:#FFFFFF !important;
    opacity:0.4 !important;
    border-left:4px solid #B8960C !important;
    cursor:not-allowed !important;
}

/* ── SECONDARY button — mode selectors, US/CA samples, Clear ─────────────── */
.stButton button[kind="secondary"],
.stButton button:not([kind="primary"]) {
    background:#FFFFFF !important;
    color:#1C2B3A !important;
    border:1.5px solid #1C2B3A !important;
    border-radius:8px !important;
    font-family:'DM Sans',sans-serif !important;
    font-weight:500 !important;
    font-size:13px !important;
    transition:background 0.15s, border-color 0.15s !important;
}
.stButton button[kind="secondary"]:hover,
.stButton button:not([kind="primary"]):hover {
    background:#F0EDE6 !important;
    border-color:#B8960C !important;
    color:#1C2B3A !important;
}

/* ── Download buttons ──────────────────────────────────────────────────────── */
.stDownloadButton > button {
    background:#FFFFFF !important;
    border:1px solid #E8E4DC !important;
    color:#1C2B3A !important;
    font-family:'DM Sans',sans-serif !important;
    font-weight:500 !important;
    font-size:13px !important;
    border-radius:8px !important;
    transition:all 0.15s !important;
}
.stDownloadButton > button:hover {
    border-color:#B8960C !important;
    color:#B8960C !important;
}

/* ── File uploader ─────────────────────────────────────────────────────────── */
[data-testid="stFileUploader"],
[data-testid="stFileUploaderDropzone"] {
    background:#FFFFFF !important;
    border:1px dashed #E8E4DC !important;
    border-radius:8px !important;
    transition:border-color 0.2s !important;
}
[data-testid="stFileUploader"]:hover,
[data-testid="stFileUploaderDropzone"]:hover {
    border-color:#B8960C !important;
    border-style:solid !important;
}
[data-testid="stFileUploaderDropzone"] button,
[data-testid="stFileUploader"] button {
    background:#1C2B3A !important;
    color:#FFFFFF !important;
    border:none !important;
    border-radius:6px !important;
    font-family:'DM Sans',sans-serif !important;
    font-size:13px !important;
}
[data-testid="stFileUploaderDropzone"] > div { color:#1C2B3A !important; }
[data-testid="stFileUploaderDropzone"] > div > small { color:#6B7280 !important; }
[data-testid="stFileUploaderDropzone"] svg { color:#B8960C !important; fill:#B8960C !important; }

/* ── Progress bar ──────────────────────────────────────────────────────────── */
.stProgress > div > div > div > div {
    background:linear-gradient(90deg,#8A7009,#B8960C) !important;
}
.stProgress > div > div {
    background:#E8E4DC !important;
    border-radius:4px !important;
}

/* ── Metrics ───────────────────────────────────────────────────────────────── */
[data-testid="metric-container"] {
    background:#FFFFFF !important;
    border:1px solid #E8E4DC !important;
    border-radius:8px !important;
    padding:1rem !important;
    box-shadow:0 1px 4px rgba(0,0,0,0.06) !important;
}
[data-testid="stMetricValue"] {
    font-family:'JetBrains Mono',monospace !important;
    font-size:1.1rem !important;
    color:#1C2B3A !important;
}
[data-testid="stMetricLabel"] {
    font-size:11px !important;
    text-transform:uppercase !important;
    letter-spacing:0.1em !important;
    color:#6B7280 !important;
}

/* ── Expander ──────────────────────────────────────────────────────────────── */
[data-testid="stExpander"] {
    background:#FFFFFF !important;
    border:1px solid #E8E4DC !important;
    border-radius:8px !important;
    box-shadow:0 1px 4px rgba(0,0,0,0.04) !important;
}
details summary p { color:#6B7280 !important; font-size:13px !important; }

/* ── Checkbox ──────────────────────────────────────────────────────────────── */
.stCheckbox label span,
label[data-testid="stWidgetLabel"] p {
    color:#6B7280 !important;
    font-size:13px !important;
    font-family:'DM Sans',sans-serif !important;
}

/* ── Slider ────────────────────────────────────────────────────────────────── */
div[data-testid="stSlider"] label p {
    color:#6B7280 !important;
    font-size:13px !important;
}

/* ── Alerts / info boxes ───────────────────────────────────────────────────── */
[data-testid="stAlert"] { border-radius:8px !important; }

/* ── Log box ───────────────────────────────────────────────────────────────── */
.run-log {
    background:#FAFAF8;
    border:1px solid #E8E4DC;
    border-left:3px solid #B8960C;
    border-radius:6px;
    padding:10px 16px;
    font-family:'JetBrains Mono',monospace;
    font-size:12px;
    color:#9CA3AF;
    max-height:160px;
    overflow-y:auto;
    line-height:1.75;
}
.run-log .ok   { color:#15803D; }
.run-log .warn { color:#B45309; }
.run-log .err  { color:#B91C1C; }
.run-log .info { color:#B8960C; }
.run-log .dim  { color:#9CA3AF; }

/* ── Divider ───────────────────────────────────────────────────────────────── */
hr { border-color:#E8E4DC !important; margin:1.5rem 0 !important; }

/* ── Dataframe ─────────────────────────────────────────────────────────────── */
[data-testid="stDataFrame"] { border-radius:8px !important; overflow:hidden !important; }
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# ── Session state (canonical keys) ────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════

if "results" not in st.session_state:
    st.session_state.results = []
if "results_mode" not in st.session_state:
    st.session_state.results_mode = "Both"
if "research_mode" not in st.session_state:
    st.session_state.research_mode = "Both"
if "elapsed_total" not in st.session_state:
    st.session_state.elapsed_total = 0
if "ticker_area" not in st.session_state:
    st.session_state.ticker_area = ""
if "spac_mode" not in st.session_state:
    st.session_state.spac_mode = False
if "running" not in st.session_state:
    st.session_state.running = False
if "nav_page" not in st.session_state:
    st.session_state.nav_page = "Ticker Research"
if "spac_enriching" not in st.session_state:
    st.session_state.spac_enriching = False
if "spac_filter" not in st.session_state:
    st.session_state.spac_filter = "All"
if "spac_search" not in st.session_state:
    st.session_state.spac_search = ""

_API_KEY  = os.getenv("SALESQL_API_KEY", "")
_API_LIVE = bool(_API_KEY)


# ══════════════════════════════════════════════════════════════════════════════
# ── Helpers ───────────────────────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════

def _parse_tickers(raw: str) -> list[str]:
    seen: list[str] = []
    for p in re.split(r"[,\s\n\r]+", (raw or "").strip()):
        p = p.strip().upper()
        if p and p not in seen:
            seen.append(p)
    return seen


def _tickers_from_file(f) -> list[str]:
    try:
        name = f.name.lower()
        if name.endswith(".xlsx"):
            df  = pd.read_excel(f)
            col = next(
                (c for c in df.columns if c.lower() in ("ticker","symbol","tickers","symbols")),
                df.columns[0],
            )
            return [str(v).strip().upper() for v in df[col].dropna() if str(v).strip()]
        raw = f.read().decode("utf-8", errors="ignore")
        return _parse_tickers(raw)
    except Exception:
        return []


def _role_label(title: str) -> str:
    t = title.lower()
    parts = re.split(r"[\s,&|/]+", t)
    if "chief executive" in t or "ceo" in parts: return "CEO"
    if "chief financial"  in t or "cfo" in parts: return "CFO"
    return title[:20]


def _fmt(val, kind: str = "") -> str:
    if val is None or val == "": return "—"
    try:
        v = float(val)
        if kind == "price":  return f"${v:,.2f}"
        if kind == "mcap":
            if v >= 1e9: return f"${v/1e9:.1f}B"
            if v >= 1e6: return f"${v/1e6:.1f}M"
            return f"${v:,.0f}"
        if kind == "cash":
            if abs(v) >= 1e9: return f"${v/1e9:.2f}B"
            if abs(v) >= 1e6: return f"${v/1e6:.1f}M"
            return f"${v:,.0f}"
        if kind == "vol":
            if v >= 1e6: return f"{v/1e6:.1f}M"
            if v >= 1e3: return f"{v/1e3:.0f}K"
            return f"{v:.0f}"
    except Exception:
        pass
    return str(val)


# ══════════════════════════════════════════════════════════════════════════════
# ── Pipeline ──────────────────────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════

_LI_TIMEOUT = 12   # seconds before we give up on LinkedIn
_EMPTY_IR   = {"ir_email": None, "ir_contact": None, "ir_page": None}


def _find_li_safe(name: str, company: str, title: str) -> str:
    """Run find_linkedin_url with a hard timeout via daemon thread."""
    result_box: list = [None]
    def _run():
        try:
            result_box[0] = find_linkedin_url(name, company, title) or ""
        except Exception:
            result_box[0] = ""
    t = threading.Thread(target=_run, daemon=True)
    t.start()
    t.join(timeout=_LI_TIMEOUT)
    return result_box[0] or ""


def _enrich_spac_executive(
    name: str, role: str, title: str, company: str, website: str,
    ticker: str, delay: float, skip_linkedin: bool,
    spac_domain_info: dict, log_fn, prog_fn,
) -> dict:
    """
    SPAC-specific 6-step fallback chain for enriching a single executive.

    STEP 1: yfinance website (already in `website` param — handled by caller)
    STEP 2: EDGAR domain (already in spac_domain_info["domain"])
    STEP 3: Name-only SalesQL search (name + company name, no domain)
    STEP 4: LinkedIn URL → SalesQL enrich by URL
    STEP 5: Sponsor domain (already in spac_domain_info["domain"] if source=sponsor)
    STEP 6: Return empty with SPAC flag
    """
    first, last = split_name(name)
    enrichment = salesql_empty("not_tried")

    # ── Attempt with EDGAR/sponsor domain (STEPS 2 & 5) ──────────────────────
    spac_domain = spac_domain_info.get("domain", "")
    if spac_domain:
        spac_source = spac_domain_info.get("source", "")
        prog_fn(f"{ticker}  —  SalesQL ({spac_source}): {role} {name} …")
        log_fn(f"   [{ticker}]  SalesQL with {spac_source} domain ({spac_domain}) → {role}: {name} …", "dim")
        time.sleep(delay)
        enrichment = search_by_name_with_variations(
            first, last, name, company, website=f"https://{spac_domain}", is_spac=True,
        )
        if enrichment.get("best_email") or enrichment.get("phone"):
            log_fn(
                f"  ✓ {role}: {name}  →  "
                f"{enrichment.get('best_email','—')}  {enrichment.get('phone','—')}  "
                f"(via {spac_source})",
                "ok",
            )
            return enrichment

    # ── STEP 3: Name-only SalesQL (name + company name, no domain) ────────────
    prog_fn(f"{ticker}  —  SPAC name search: {role} {name} …")
    log_fn(f"   [{ticker}]  SPAC STEP 3: name-only SalesQL → {role}: {name} …", "dim")
    time.sleep(delay)
    enrichment = search_by_name_and_company(name, company, title=role)
    if enrichment.get("best_email") or enrichment.get("phone"):
        log_fn(
            f"  ✓ {role}: {name}  →  "
            f"{enrichment.get('best_email','—')}  {enrichment.get('phone','—')}  "
            f"(name-only search)",
            "ok",
        )
        return enrichment

    # ── STEP 4: LinkedIn URL → SalesQL enrich by URL ──────────────────────────
    if not skip_linkedin:
        prog_fn(f"{ticker}  —  SPAC LinkedIn: {role} {name} …")
        log_fn(f"   [{ticker}]  SPAC STEP 4: LinkedIn search → {role}: {name} …", "dim")
        li_url = _find_li_safe(name, company, title)
        if li_url:
            log_fn(f"   [{ticker}]  LinkedIn found: {li_url}  →  SalesQL enriching …", "dim")
            prog_fn(f"{ticker}  —  SalesQL via LinkedIn: {role} {name} …")
            time.sleep(delay)
            enrichment = enrich_by_url(li_url)
            if enrichment.get("best_email") or enrichment.get("phone"):
                enrichment["linkedin_url"] = li_url
                log_fn(
                    f"  ✓ {role}: {name}  →  "
                    f"{enrichment.get('best_email','—')}  {enrichment.get('phone','—')}  "
                    f"(via LinkedIn)",
                    "ok",
                )
                return enrichment
        else:
            log_fn(f"   [{ticker}]  LinkedIn: not found for {name}", "dim")

    # ── Also try standard domain-based variations with is_spac=True ───────────
    if website:
        prog_fn(f"{ticker}  —  SalesQL standard: {role} {name} …")
        log_fn(f"   [{ticker}]  SPAC fallback: standard SalesQL with website domain …", "dim")
        time.sleep(delay)
        enrichment = search_by_name_with_variations(
            first, last, name, company, website=website, is_spac=True,
        )
        if enrichment.get("best_email") or enrichment.get("phone"):
            return enrichment

    # ── STEP 6: Return what we have (empty enrichment, but name is populated) ──
    log_fn(f"  ⚠ {role}: {name}  →  no contact found (SPAC — all fallbacks exhausted)", "warn")
    return salesql_empty("spac_not_found")


def process_ticker(ticker: str, log_fn, prog_fn, skip_linkedin: bool, delay: float,
                   mode: str = "Both", run_bouncer: bool = True,
                   is_spac: bool = False) -> dict:
    """
    Full enrichment pipeline for one ticker.
    mode: "Both" | "Financial Analysis Only" | "Person Lookup Only"
    is_spac: when True (or auto-detected), activates SPAC fallback chain
    """
    ticker = ticker.strip().upper()
    log_fn(f"▶  [{ticker}]  resolving company …", "dim")
    prog_fn(f"{ticker}  —  resolving …")

    try:
        info = resolve_ticker(ticker)
    except Exception as e:
        log_fn(f"✗  [{ticker}]  resolve error: {e}", "err")
        return empty_row(ticker, "API Error")

    if not info:
        log_fn(f"✗  [{ticker}]  not recognised — check spelling (e.g. AAPL not APPL)", "err")
        return empty_row(ticker, "Not Found")

    company = info.get("company") or ""
    website = info.get("website") or ""
    targets = info.get("targets") or []

    if not company:
        log_fn(f"⚠  [{ticker}]  no company data returned — skipping", "warn")
        return empty_row(ticker, "No Data")

    # ── SPAC auto-detection ─────────────────────────────────────────────────────
    if not is_spac:
        try:
            import yfinance as yf
            yf_info = yf.Ticker(ticker).info or {}
            is_spac = detect_spac(yf_info, ticker)
        except Exception:
            pass

    if is_spac:
        log_fn(f"◈  [{ticker}]  Detected as SPAC — activating SPAC fallback lookup chain", "info")
        info["is_spac"] = True

    log_fn(f"✓  [{ticker}]  {company}  ·  {len(targets)} exec(s)", "ok")

    if not targets and mode != "Financial Analysis Only":
        # Fallback 1: Web search for "<Company> CEO/CFO" names
        log_fn(f"⚠  [{ticker}]  no executives listed — searching web for CEO/CFO …", "warn")
        try:
            from lookup.web_search_fallback import search_executives as _web_search
            prog_fn(f"{ticker}  —  web search: CEO/CFO names …")
            _web_hits = _web_search(company, roles=["CEO", "CFO"], delay=delay)
            for _wh in _web_hits:
                targets.append(_wh)
                log_fn(f"  ✓ Found {_wh['title']} via web search: {_wh['name']}", "ok")
        except Exception:
            pass

        # Fallback 2: SalesQL title search (if web search still didn't find anyone)
        if not targets:
            log_fn(f"   [{ticker}]  web search empty — trying SalesQL title search …", "dim")
            for _fb_role, _fb_title in [("CEO", "CEO"), ("CFO", "CFO")]:
                try:
                    prog_fn(f"{ticker}  —  SalesQL title search: {_fb_role} …")
                    time.sleep(delay)
                    _fb_result = search_by_name_and_company("", company, title=_fb_title)
                    _fb_name = (_fb_result.get("full_name") or "").strip()
                    if _fb_name and (_fb_result.get("best_email") or _fb_result.get("phone")):
                        targets.append({"name": _fb_name, "title": _fb_title})
                        log_fn(f"  ✓ Found {_fb_role} via title search: {_fb_name}", "ok")
                except Exception:
                    pass

    # ── Financials ─────────────────────────────────────────────────────────────
    fin: dict = {}
    if mode in ("Both", "Financial Analysis Only"):
        prog_fn(f"{ticker}  —  fetching financials …")
        log_fn(f"   [{ticker}]  fetching financials …", "dim")
        try:
            fin = fetch_financials_safe(ticker)
            log_fn(
                f"   price={_fmt(fin.get('stock_price'),'price')}  "
                f"mcap={_fmt(fin.get('market_cap'),'mcap')}",
                "dim",
            )
        except Exception:
            fin = {}

    # Short-circuit for Financial Analysis Only
    if mode == "Financial Analysis Only":
        try:
            import yfinance as yf
            info["exchange"] = (yf.Ticker(ticker).info or {}).get("exchange", "")
        except Exception:
            pass
        row = build_row(ticker=ticker, company_info=info, financials=fin,
                        executives=[], ir_data={})
        log_fn(f"✅  [{ticker}]  complete (financials only)", "ok")
        prog_fn(f"{ticker}  —  done ✓")
        return row

    # ── SPAC domain resolution (STEPS 2 & 5 — run once, shared by all execs) ──
    spac_domain_info: dict = {}
    if is_spac and not website:
        prog_fn(f"{ticker}  —  SPAC: resolving domain via SEC EDGAR …")
        try:
            spac_domain_info = resolve_spac_domain(ticker, company, log_fn=log_fn)
            spac_dom = spac_domain_info.get("domain", "")
            if spac_dom:
                website = f"https://{spac_dom}"
                log_fn(f"   [{ticker}]  SPAC domain resolved: {spac_dom} (via {spac_domain_info.get('source','')})", "info")
        except Exception as e:
            log_fn(f"   [{ticker}]  SPAC EDGAR lookup failed: {e}", "warn")
            spac_domain_info = {}

    # ── Executives ─────────────────────────────────────────────────────────────
    executives: list[dict] = []
    ceo_t = next((t for t in targets if _role_label(t["title"]) == "CEO"), None)
    cfo_t = next((t for t in targets if _role_label(t["title"]) == "CFO"), None)
    if not ceo_t and targets:
        ceo_t = targets[0]

    for target in [ceo_t, cfo_t]:
        if not target:
            continue
        name  = re.sub(r"^\s*(Mr\.?|Ms\.?|Mrs\.?|Dr\.?|Prof\.?)\s+", "", target["name"], flags=re.IGNORECASE).strip()
        name  = re.sub(r"\s+", " ", name).strip()
        title = target["title"]
        role  = _role_label(title)
        first, last = split_name(name)

        try:
            # ── SPAC PATH: full 6-step fallback chain ─────────────────────────
            if is_spac:
                enrichment = _enrich_spac_executive(
                    name=name, role=role, title=title, company=company,
                    website=website, ticker=ticker, delay=delay,
                    skip_linkedin=skip_linkedin,
                    spac_domain_info=spac_domain_info,
                    log_fn=log_fn, prog_fn=prog_fn,
                )
            else:
                # ── NORMAL PATH ────────────────────────────────────────────────
                enrichment = salesql_empty("not_tried")

                if not skip_linkedin:
                    prog_fn(f"{ticker}  —  LinkedIn: {role} {name} …")
                    log_fn(f"   [{ticker}]  LinkedIn → {role}: {name} …", "dim")
                    li_url = _find_li_safe(name, company, title)
                    if li_url:
                        log_fn(f"   [{ticker}]  LinkedIn found  →  SalesQL enriching …", "dim")
                        prog_fn(f"{ticker}  —  SalesQL: {role} {name} …")
                        time.sleep(delay)
                        enrichment = enrich_by_url(li_url)
                        if enrichment.get("best_email") or enrichment.get("phone"):
                            enrichment["linkedin_url"] = li_url
                            log_fn(
                                f"  ✓ {role}: {name}  →  "
                                f"{enrichment.get('best_email','—')}  {enrichment.get('phone','—')}",
                                "ok",
                            )
                            executives.append({"role": role, "name": name, "title": title, "enrichment": enrichment})
                            continue
                    else:
                        log_fn(f"   [{ticker}]  LinkedIn: not found — falling back to SalesQL name search", "dim")

                prog_fn(f"{ticker}  —  SalesQL name search: {role} {name} …")
                log_fn(f"   [{ticker}]  SalesQL name search → {role}: {name} …", "dim")
                time.sleep(delay)
                enrichment = search_by_name_with_variations(first, last, name, company, website=website, is_spac=is_spac)

                # ── Auto LinkedIn fallback: if SalesQL found nothing, try LinkedIn
                #    even when the global toggle is off ─────────────────────────────
                if not enrichment.get("best_email") and not enrichment.get("phone") and skip_linkedin:
                    log_fn(f"   [{ticker}]  SalesQL empty → auto LinkedIn fallback for {name} …", "dim")
                    prog_fn(f"{ticker}  —  LinkedIn fallback: {role} {name} …")
                    li_url = _find_li_safe(name, company, title)
                    if li_url:
                        log_fn(f"   [{ticker}]  LinkedIn found: {li_url} → enriching …", "dim")
                        time.sleep(delay)
                        li_enrichment = enrich_by_url(li_url)
                        if li_enrichment.get("best_email") or li_enrichment.get("phone"):
                            li_enrichment["linkedin_url"] = li_url
                            enrichment = li_enrichment
                            log_fn(f"  ✓ {role}: {name} → {enrichment.get('best_email','—')} (via LinkedIn fallback)", "ok")

                em = enrichment.get("best_email") or "—"
                ph = enrichment.get("phone") or "—"
                log_fn(
                    f"  {'✓' if em != '—' else '⚠'} {role}: {name}  →  {em}  {ph}",
                    "ok" if em != "—" else "warn",
                )

        except Exception as e:
            log_fn(f"  ✗ {role} error: {e}", "err")
            enrichment = salesql_empty("error")

        executives.append({"role": role, "name": name, "title": title, "enrichment": enrichment})

    # ── Cross-domain retry ─────────────────────────────────────────────────────
    try:
        from lookup.salesql_enricher import _extract_domain
        from lookup.schema_builder   import _is_personal_email
        site_dom   = _extract_domain(website) if website else ""
        brand_doms: set[str] = set()
        for ex in executives:
            em = (ex["enrichment"].get("best_email") or "").strip()
            if em and "@" in em:
                d = em.split("@")[-1].lower()
                if d != site_dom and not _is_personal_email(em) and not d.endswith(".edu") and len(d) > 4:
                    brand_doms.add(d)
        if brand_doms:
            for ex in executives:
                best = (ex["enrichment"].get("best_email") or "").strip()
                if best and not _is_personal_email(best) and not best.endswith(".edu"):
                    continue
                f2, l2 = split_name(ex["name"])
                for bd in brand_doms:
                    time.sleep(delay)
                    retry = search_by_name_with_variations(
                        f2, l2, ex["name"], company, website=f"https://{bd}",
                        is_spac=is_spac,
                    )
                    rm = (retry.get("best_email") or "").strip()
                    if rm and "@" in rm and not _is_personal_email(rm) and not rm.endswith(".edu"):
                        ex["enrichment"] = retry
                        log_fn(f"  ↺ {ex['name']}  →  {rm}", "info")
                        break
    except Exception:
        pass

    # ── Email pattern fill (domain-aware inference) ───────────────────────────
    try:
        from lookup.email_pattern import detect_pattern
        from lookup.schema_builder import _is_personal_email

        # Extract company domain for validation
        comp_domain = ""
        if website:
            try:
                from lookup.salesql_enricher import _extract_domain
                comp_domain = _extract_domain(website)
            except Exception:
                comp_domain = re.sub(r"https?://(www\.)?", "", website).split("/")[0]

        def _valid_co_email(em: str) -> bool:
            if not em or "@" not in em:
                return False
            if _is_personal_email(em):
                return False
            if not comp_domain:
                return True
            em_dom   = em.split("@")[-1].lower().lstrip("www.")
            corp_dom = comp_domain.lower().lstrip("www.")
            corp_dom = corp_dom.replace("https://", "").replace("http://", "").split("/")[0]
            return (em_dom == corp_dom
                    or em_dom.endswith("." + corp_dom)
                    or corp_dom.endswith("." + em_dom))

        ep_list = []
        for e in executives:
            raw_em = (e["enrichment"].get("best_email") or "").strip()
            ep_list.append({
                "name":       e["name"],
                "first_name": split_name(e["name"])[0],
                "last_name":  split_name(e["name"])[1],
                "best_email": raw_em if _valid_co_email(raw_em) else "",
            })

        pattern = detect_pattern(ep_list, website=website)

        if pattern:
            for e, ep in zip(executives, ep_list):
                if ep["best_email"]:
                    continue
                first = ep["first_name"]
                last  = ep["last_name"]
                guess = pattern.guess(first, last)
                if guess:
                    e["enrichment"]["best_email"] = guess
                    e["enrichment"]["work_email"]  = guess
                    e["enrichment"]["is_inferred"] = True
                    log_fn(
                        f"   [{ticker}]  ✦ inference: {e['name']} → {guess}"
                        f"  (pattern: {pattern.name} · source: {pattern.domain})",
                        "info",
                    )
        else:
            fill_missing_emails(ep_list, website=website, verbose=False)
            for e, ep in zip(executives, ep_list):
                if not e["enrichment"].get("best_email") and ep.get("best_email"):
                    e["enrichment"]["best_email"] = ep["best_email"]
                    e["enrichment"]["work_email"]  = ep["best_email"]

    except Exception:
        pass

    # ── IR data (hard 20s timeout — IR probing can hang on slow sites) ────────
    prog_fn(f"{ticker}  —  finding IR contact …")
    log_fn(f"   [{ticker}]  fetching IR data …", "dim")
    try:
        _ir_box: list = [dict(_EMPTY_IR)]
        def _ir_run():
            try:
                _ir_box[0] = find_ir_data(ticker=ticker, company=company, website=website)
            except Exception:
                pass
        _ir_t = threading.Thread(target=_ir_run, daemon=True)
        _ir_t.start()
        _ir_t.join(timeout=20)
        ir = _ir_box[0]
        log_fn(
            f"  {'✓' if ir.get('ir_email') else '·'} IR  "
            f"email={ir.get('ir_email') or '—'}  page={ir.get('ir_page') or '—'}",
            "ok" if ir.get("ir_email") else "dim",
        )
        ir_contact = ir.get("ir_contact", "")
        if ir_contact and not ir.get("ir_email"):
            plain = re.sub(r"\s*\(.*\)\s*$", "", ir_contact).split(",")[0].strip()
            _bad  = {"firm","group","associates","partners","company","corp","inc","llc","ltd",
                     "department","dept","team","division","office","relations","contact",
                     "capital","management","services","solutions","communications"}
            f3, l3 = split_name(plain)
            if f3 and l3 and l3.lower().rstrip(".,") not in _bad:
                try:
                    ir_en = search_by_name_with_variations(f3, l3, plain, company, website=website, is_spac=is_spac)
                    em = ir_en.get("best_email") or ir_en.get("work_email") or ""
                    if em:
                        ir["ir_email"] = em
                except Exception:
                    pass
    except Exception:
        ir = {}

    # ── STEP 6: SPAC fallback — populate IR with EDGAR link ───────────────────
    if is_spac:
        if not ir.get("ir_contact"):
            ir["ir_contact"] = "SPAC — contact via SEC filing"
        if not ir.get("ir_page"):
            edgar_url = spac_domain_info.get("edgar_url", "") if spac_domain_info else ""
            if not edgar_url:
                edgar_url = get_edgar_filing_url(ticker)
            if edgar_url:
                ir["ir_page"] = edgar_url

    # ── Exchange ───────────────────────────────────────────────────────────────
    try:
        import yfinance as yf
        info["exchange"] = (yf.Ticker(ticker).info or {}).get("exchange", "")
    except Exception:
        pass

    # ── Extract company domain for email flagging ──────────────────────────────
    comp_domain = ""
    if website:
        try:
            from lookup.salesql_enricher import _extract_domain
            comp_domain = _extract_domain(website)
        except Exception:
            comp_domain = re.sub(r"https?://(www\.)?", "", website).split("/")[0]

    row = build_row(ticker=ticker, company_info=info, financials=fin,
                    executives=executives, ir_data=ir,
                    company_domain=comp_domain)

    # ── Mark SPAC in row so UI card can display [SPAC] tag ────────────────────
    if is_spac:
        row["_is_spac"] = True

    # ── Tag inferred emails in the output row ──────────────────────────────────
    for _e in executives:
        if _e["enrichment"].get("is_inferred"):
            _role = _e.get("role", "")
            _col  = f"{_role} EMAIL"
            if _col in row:
                _em = str(row.get(_col) or "")
                _blank = {"not found", "—", "", "not on linkedin"}
                if _em.strip().lower() not in _blank and "✦ inferred" not in _em:
                    row[_col] = _em + " ✦ inferred"

    # ── Bouncer email verification ─────────────────────────────────────────────
    if run_bouncer and mode != "Financial Analysis Only":
        prog_fn(f"{ticker}  —  verifying emails with Bouncer …")
        log_fn(f"   [{ticker}]  Bouncer email verification …", "dim")
        try:
            verify_row_emails(row, delay=0.3, log_fn=log_fn)
        except Exception as e:
            log_fn(f"   [{ticker}]  Bouncer skipped: {e}", "warn")

    log_fn(f"✅  [{ticker}]  complete", "ok")
    prog_fn(f"{ticker}  —  done ✓")
    return row


# ══════════════════════════════════════════════════════════════════════════════
# ── Excel export ──────────────────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════

def build_excel(rows: list[dict], mode: str = "Both") -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils  import get_column_letter
    except ImportError:
        buf = io.BytesIO()
        pd.DataFrame(rows, columns=COLUMN_ORDER).to_excel(buf, index=False)
        return buf.getvalue()

    CONTACT_COLS  = ["Ticker","Company Name","CEO","CFO","CEO EMAIL","CEO NUMBER",
                     "CFO EMAIL","CFO NUMBER","IR Email","IR Contact","IR Page"]
    FINANCIAL_COLS = ["Ticker","Company Name","Industry","Exchange",
                      "Stock Price (Most Recent)","Market Cap (Most Recent)",
                      "Cash (Latest K)","Cash (Latest Q)","1M Share Volume",
                      "1D $ Share Volume","Cash from Ops (Latest K)","Cash from Ops (Latest Q)"]

    GOLD   = "C9A84C"
    DARK   = "0A0C10"
    ROW_A  = "111318"
    ROW_B  = "161A22"
    WHITE  = "F0EDE8"
    MUTED  = "4A4D56"
    GREEN  = "2ECC71"
    ORANGE = "F39C12"
    RED    = "E74C3C"
    BLANK  = {"","not found","not on linkedin","not on sql","n/a","—","api error","error"}

    def hdr(ws, cols):
        for ci, col in enumerate(cols, 1):
            c = ws.cell(row=1, column=ci, value=col)
            c.font      = Font(bold=True, color=DARK, size=9, name="Calibri")
            c.fill      = PatternFill("solid", fgColor=GOLD)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    def data(ws, rows, cols):
        WIDTHS = {
            "Company Name":28,"Ticker":10,"Industry":18,"Exchange":9,
            "Stock Price (Most Recent)":14,"Market Cap (Most Recent)":15,
            "Cash (Latest K)":14,"Cash (Latest Q)":14,
            "1M Share Volume":14,"1D $ Share Volume":15,
            "Cash from Ops (Latest K)":18,"Cash from Ops (Latest Q)":18,
            "CEO":26,"CFO":26,"CEO EMAIL":34,"CEO NUMBER":22,
            "CFO EMAIL":34,"CFO NUMBER":22,"IR Email":30,"IR Contact":28,"IR Page":38,
        }
        for ri, row in enumerate(rows, 2):
            bg = ROW_A if ri % 2 == 0 else ROW_B
            for ci, col in enumerate(cols, 1):
                val = row.get(col, "")
                c   = ws.cell(row=ri, column=ci, value=str(val) if val is not None else "")
                c.fill      = PatternFill("solid", fgColor=bg)
                c.alignment = Alignment(vertical="top", wrap_text=True)
                v = str(val or "").strip().lower()
                if col in ("CEO EMAIL","CFO EMAIL","IR Email"):
                    if v in BLANK:
                        c.font = Font(color=MUTED, size=9, italic=True, name="Courier New")
                    elif "⚠ invalid" in v:
                        c.font = Font(color="E74C3C", size=9, name="Courier New", strike=True)
                    elif "⚠ risky" in v:
                        c.font = Font(color="E67E22", size=9, name="Courier New")
                    elif "⚠ unverified" in v:
                        c.font = Font(color="BDC300", size=9, name="Courier New")
                    elif "✦ inferred" in v:
                        c.font = Font(color=ORANGE, size=9, name="Courier New", italic=True)
                        c.fill = PatternFill("solid", fgColor="3D2200")
                    elif "no work email found" in v or "(no work" in v:
                        c.font = Font(color=ORANGE, size=9, name="Courier New")
                    else:
                        c.font = Font(color=WHITE, size=9, name="Courier New")
                elif col in ("CEO NUMBER","CFO NUMBER"):
                    if v in BLANK:  c.font = Font(color=MUTED, size=9, italic=True)
                    elif v.startswith("work"): c.font = Font(color=GREEN, size=9, name="Courier New")
                    else:           c.font = Font(color=WHITE, size=9, name="Courier New")
                else:
                    c.font = Font(color=MUTED if v in BLANK else WHITE, size=9)
        for ci, col in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(ci)].width = WIDTHS.get(col, 16)
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = GOLD

    wb  = openpyxl.Workbook()
    ws1 = wb.active

    if mode == "Financial Analysis Only":
        ws1.title = "Financials"
        hdr(ws1, FINANCIAL_COLS); data(ws1, rows, FINANCIAL_COLS)
    elif mode == "Person Lookup Only":
        ws1.title = "Contacts"
        hdr(ws1, CONTACT_COLS); data(ws1, rows, CONTACT_COLS)
    else:
        ws1.title = "Full Output"
        ws2 = wb.create_sheet("Contacts Only")
        ws3 = wb.create_sheet("Financials Only")
        hdr(ws1, COLUMN_ORDER);   data(ws1, rows, COLUMN_ORDER)
        hdr(ws2, CONTACT_COLS);   data(ws2, rows, CONTACT_COLS)
        hdr(ws3, FINANCIAL_COLS); data(ws3, rows, FINANCIAL_COLS)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# ── Display helpers ───────────────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════

_BLANK_DISP = {
    "", "not found", "not on linkedin", "not on sql", "n/a", "—",
    "api error", "error", "not found on linkedin", "no data",
}

def _display_email(val: str) -> tuple[str, str]:
    v  = str(val or "").strip()
    vl = v.lower()
    if vl in _BLANK_DISP:
        return "—", "#D1CEC8"
    if "⚠ invalid" in vl:
        return "—", "#D1CEC8"
    clean = re.sub(r"\s*⚠\s+\w+", "", v).strip()
    clean = re.sub(r"\s*\(no mx\)|\s*\(disposable\)|\s*\(role acct\)", "", clean).strip()
    if "no work email found" in vl or "(no work" in vl:
        bare = clean.split("(")[0].strip()
        return bare or "—", "#B45309"
    if "✦ inferred" in clean:
        return clean, "#B8960C"
    return clean or v, "#1C2B3A"


def _display_phone(val: str) -> tuple[str, str]:
    v  = str(val or "").strip()
    vl = v.lower()
    if vl in _BLANK_DISP:
        return "—", "#D1CEC8"
    color = "#15803D" if vl.startswith("work") else "#1C2B3A"
    return v, color


def _clean_ir_contact(contact: str) -> str:
    if not contact or str(contact).strip().lower() in _BLANK_DISP:
        return "—"
    c = str(contact).strip()
    if len(c) > 80:
        return "—"
    _nav = {
        "sec","filings","governance","corporate","management","committee",
        "documents","presentations","annual","quarterly","report","overview",
        "releases","events","calendar","press","news","ki","menu","nav",
    }
    words = set(re.findall(r'[a-z]+', c.lower()))
    if len(words & _nav) >= 2:
        return "—"
    if not re.match(r'^[A-Z]', c):
        return "—"
    return c


def _count_populated(rows: list[dict]) -> tuple[int, int]:
    cols = ["CEO EMAIL","CEO NUMBER","CFO EMAIL","CFO NUMBER","IR Email","IR Contact"]
    total = missing = 0
    for row in rows:
        for col in cols:
            total += 1
            v = str(row.get(col, "") or "").strip().lower()
            if v in _BLANK_DISP or v == "not found":
                missing += 1
    return total, missing


# ══════════════════════════════════════════════════════════════════════════════
# ── Company card renderer ─────────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════

_LOGO_SVG = (
    '<svg viewBox="0 0 370 68" xmlns="http://www.w3.org/2000/svg" '
    'style="height:38px;width:auto;display:block;">'
    '<text x="187" y="52" font-family="Georgia,\'Times New Roman\',serif" '
    'font-size="42" fill="#1A2332" letter-spacing="9" text-anchor="middle" '
    'font-weight="400">CURVATURE</text>'
    '<path d="M 156 38 Q 188 5 220 38" fill="none" stroke="#B8960C" '
    'stroke-width="3.5" stroke-linecap="round"/>'
    '</svg>'
)


def _company_card_html(row: dict, mode: str) -> str:
    """Build the full HTML card for one company."""
    ticker   = str(row.get("Ticker", "") or "").strip()
    company  = str(row.get("Company Name", "") or "").strip()
    exchange = str(row.get("Exchange", "") or "").strip()
    industry = (row.get("Industry") or "—")[:26]

    is_error = company.lower() in ("not found","api error","no data","error")
    tick_col = "#B45309" if is_error else "#B8960C"
    warn     = (' <span style="color:#B45309;font-size:12px;" '
                'title="No data found — verify this ticker symbol is correct">⚠ verify ticker</span>'
                if is_error else "")
    spac_tag = (' <span style="background:#B8960C18;border:1px solid #B8960C55;color:#B8960C;'
                'font-size:10px;font-weight:700;letter-spacing:0.08em;padding:2px 7px;'
                'border-radius:4px;margin-left:6px;vertical-align:middle;">SPAC</span>'
                if row.get("_is_spac") else "")

    # Card wrapper + header
    html = (
        f'<div style="border:1px solid #E8E4DC;border-radius:8px;'
        f'margin-bottom:16px;overflow:hidden;'
        f'box-shadow:0 1px 4px rgba(0,0,0,0.06);">'
        f'<div style="background:#FFFFFF;padding:12px 20px;display:flex;'
        f'align-items:center;gap:12px;border-bottom:1px solid #E8E4DC;">'
        f'<span style="font-family:JetBrains Mono,monospace;font-size:15px;'
        f'font-weight:700;color:{tick_col};">{ticker}</span>'
        f'<span style="font-size:14px;font-weight:600;color:#1C2B3A;">'
        f'{company}{warn}{spac_tag}</span>'
        f'<span style="margin-left:auto;font-size:11px;color:#9CA3AF;'
        f'border:1px solid #E8E4DC;border-radius:4px;padding:2px 8px;'
        f'letter-spacing:0.06em;background:#FAFAF8;">{exchange}</span>'
        f'</div>'
    )

    # Financials strip
    if mode in ("Both", "Financial Analysis Only"):
        fin_items = [
            ("Price",      _fmt(row.get("Stock Price (Most Recent)"), "price"), "Current share price"),
            ("Mkt Cap",    _fmt(row.get("Market Cap (Most Recent)"),  "mcap"),  "Total market value of all shares"),
            ("Cash (Q)",   _fmt(row.get("Cash (Latest Q)"),           "cash"),  "Cash on hand — most recent quarterly filing"),
            ("Cash (K)",   _fmt(row.get("Cash (Latest K)"),           "cash"),  "Cash on hand — most recent annual filing"),
            ("Ops CF (Q)", _fmt(row.get("Cash from Ops (Latest Q)"), "cash"),   "Cash from operations last quarter"),
            ("Ops CF (K)", _fmt(row.get("Cash from Ops (Latest K)"), "cash"),   "Cash from operations last year"),
            ("1M Vol",     _fmt(row.get("1M Share Volume"),           "vol"),   "Total shares traded over the past month"),
            ("1D $ Vol",   _fmt(row.get("1D $ Share Volume"),         "vol"),   "Dollar value of shares traded yesterday"),
            ("Industry",   industry,                                             "Sector / industry classification"),
        ]
        cells = "".join(
            f'<div style="padding:10px 14px;border-right:1px solid #E8E4DC;'
            f'flex:1;min-width:76px;" title="{tip}">'
            f'<div style="font-size:10px;color:#9CA3AF;text-transform:uppercase;'
            f'letter-spacing:0.09em;margin-bottom:4px;font-family:DM Sans,sans-serif;">{lbl}</div>'
            f'<div style="font-family:JetBrains Mono,monospace;font-size:13px;'
            f'color:#1C2B3A;font-weight:500;">{val}</div>'
            f'</div>'
            for lbl, val, tip in fin_items
        )
        html += (
            f'<div style="background:#FAFAF8;display:flex;flex-wrap:wrap;'
            f'border-bottom:1px solid #E8E4DC;">{cells}</div>'
        )

    # Contacts strip
    if mode in ("Both", "Person Lookup Only"):
        def _person(role: str, name: str, email_raw: str, phone_raw: str) -> str:
            em, ec = _display_email(email_raw)
            ph, pc = _display_phone(phone_raw)
            nm = name if name and name.lower() not in _BLANK_DISP else "—"
            return (
                f'<div style="flex:1;min-width:165px;padding:14px 18px;'
                f'border-right:1px solid #E8E4DC;">'
                f'<div style="font-size:10px;color:#B8960C;text-transform:uppercase;'
                f'letter-spacing:0.1em;font-weight:600;margin-bottom:6px;">{role}</div>'
                f'<div style="font-size:14px;color:#1C2B3A;margin-bottom:5px;font-weight:500;">{nm}</div>'
                f'<div style="font-family:JetBrains Mono,monospace;font-size:12px;'
                f'color:{ec};margin-bottom:3px;" '
                f'title="{role} EMAIL — navy=verified, gold=inferred, amber=no work email, grey=not found">{em}</div>'
                f'<div style="font-family:JetBrains Mono,monospace;font-size:12px;'
                f'color:{pc};" title="Dark green = work/office number · Navy = mobile">{ph}</div>'
                f'</div>'
            )

        ir_em, ir_ec = _display_email(str(row.get("IR Email", "") or ""))
        ir_ct = _clean_ir_contact(str(row.get("IR Contact", "") or ""))
        ir_pg = str(row.get("IR Page", "") or "").strip()
        ir_link = (
            f'<a href="{ir_pg}" target="_blank" style="color:#B8960C;font-size:12px;'
            f'text-decoration:none;display:inline-block;margin-top:4px;">'
            f'↗ IR Page</a>'
            if ir_pg.startswith("http") else ""
        )
        ir_block = (
            f'<div style="flex:1;min-width:165px;padding:14px 18px;">'
            f'<div style="font-size:10px;color:#B8960C;text-transform:uppercase;'
            f'letter-spacing:0.1em;font-weight:600;margin-bottom:6px;">IR</div>'
            f'<div style="font-family:JetBrains Mono,monospace;font-size:12px;'
            f'color:{ir_ec};margin-bottom:4px;" title="Investor Relations email">{ir_em}</div>'
            f'<div style="font-size:13px;color:#6B7280;margin-bottom:0;" '
            f'title="IR Contact — named investor relations representative">{ir_ct}</div>'
            f'{ir_link}'
            f'</div>'
        )

        ceo_b = _person("CEO",
                        str(row.get("CEO","") or ""),
                        str(row.get("CEO EMAIL","") or ""),
                        str(row.get("CEO NUMBER","") or ""))
        cfo_b = _person("CFO",
                        str(row.get("CFO","") or ""),
                        str(row.get("CFO EMAIL","") or ""),
                        str(row.get("CFO NUMBER","") or ""))

        html += (
            f'<div style="background:#FFFFFF;display:flex;flex-wrap:wrap;">'
            f'{ceo_b}{cfo_b}{ir_block}'
            f'</div>'
        )

    html += '</div>'
    return html


# ══════════════════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════════════════
# ── U I   L A Y O U T  ──────────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════════════════

# ── 1. Sidebar (advanced settings) ────────────────────────────────────────────
with st.sidebar:
    st.markdown("**API Status**")
    if _API_LIVE:
        st.markdown(f"🟢 SalesQL connected  \n`…{_API_KEY[-6:]}`")
    else:
        st.markdown("🔴 SalesQL key missing  \nAdd `SALESQL_API_KEY` to `.env`")
    st.markdown("🟢 Yahoo Finance  🟢 Bouncer")
    st.divider()
    st.markdown("**Advanced Settings**")
    run_bouncer = st.toggle(
        "Verify emails with Bouncer",
        value=True,
        help="Calls Bouncer API to check deliverability. ~3 credits per ticker. "
             "Invalid addresses are replaced with — in results.",
    )
    skip_li_raw = st.toggle(
        "Enable LinkedIn lookup",
        value=False,
        help="For regular ticker lookups — adds LinkedIn profile search. "
             "Slower but finds more contacts. "
             "SPAC Research tab always uses LinkedIn automatically.",
    )
    skip_li = not skip_li_raw
    delay = st.slider(
        "SalesQL delay (s)", 0.5, 5.0, 1.5, 0.5,
        help="Seconds to wait between SalesQL API calls. Increase if you see rate-limit warnings.",
    )
    st.divider()
    st.markdown(
        '<div style="font-size:12px;color:#9CA3AF;line-height:2.2;'
        'font-family:DM Sans,sans-serif;">'
        '<b style="color:#6B7280;font-size:12px;">Email colour guide</b><br>'
        '<span style="color:#1C2B3A;">●</span> Navy — verified work email<br>'
        '<span style="color:#B8960C;">●</span> Gold — pattern-inferred<br>'
        '<span style="color:#B45309;">●</span> Amber — no work email found<br>'
        '<span style="color:#D1CEC8;">—</span> Dash — invalid or not found'
        '</div>',
        unsafe_allow_html=True,
    )


# ── 2. Header ─────────────────────────────────────────────────────────────────
st.markdown(
    f'<div style="display:flex;align-items:center;justify-content:space-between;'
    f'background:#FFFFFF;border-bottom:1px solid #E8E4DC;'
    f'padding:16px 0 16px 0;margin-bottom:24px;">'
    f'<div style="display:flex;align-items:center;gap:20px;">'
    f'{_LOGO_SVG}'
    f'<div style="border-left:1px solid #E8E4DC;padding-left:20px;">'
    f'<div style="font-size:11px;color:#6B7280;letter-spacing:0.2em;'
    f'text-transform:uppercase;font-family:DM Sans,sans-serif;font-variant:small-caps;">'
    f'Investment Banking Intelligence</div>'
    f'</div></div>'
    f'<div style="font-size:12px;color:#6B7280;font-family:DM Sans,sans-serif;">'
    f'<span style="display:inline-block;width:8px;height:8px;border-radius:50%;'
    f'background:#15803D;margin-right:6px;vertical-align:middle;"></span>LIVE'
    f'</div></div>',
    unsafe_allow_html=True,
)

# ── 2b. Navigation bar ──────────────────────────────────────────────────────
_nav_cols = st.columns([1.2, 1.3, 1.2, 4.8])
with _nav_cols[0]:
    if st.button(
        "◈  Ticker Research",
        key="nav_ticker",
        type="primary" if st.session_state.nav_page == "Ticker Research" else "secondary",
        use_container_width=True,
    ):
        st.session_state.nav_page = "Ticker Research"
        st.rerun()
with _nav_cols[1]:
    if st.button(
        "📋  SPAC Research",
        key="nav_spac",
        type="primary" if st.session_state.nav_page == "SPAC Research" else "secondary",
        use_container_width=True,
    ):
        st.session_state.nav_page = "SPAC Research"
        st.rerun()
with _nav_cols[2]:
    if st.button(
        "📤  Bulk Enrich",
        key="nav_bulk",
        type="primary" if st.session_state.nav_page == "Bulk Enrich" else "secondary",
        use_container_width=True,
    ):
        st.session_state.nav_page = "Bulk Enrich"
        st.rerun()

st.markdown("<hr style='margin:0.5rem 0 1rem 0;'>", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# ── PAGE: Ticker Research ────────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════

if st.session_state.nav_page == "Ticker Research":

    # ── 3. Welcome line ─────────────────────────────────────────────────────────
    st.markdown(
        '<p style="font-size:14px;color:#6B7280;line-height:1.75;margin-bottom:24px;'
        'font-family:DM Sans,sans-serif;">'
        'This tool automatically pulls '
        '<span style="color:#B8960C;text-decoration:underline;">financial data</span>'
        ' and '
        '<span style="color:#B8960C;text-decoration:underline;">executive contact information</span>'
        ' for any public company. '
        'Enter tickers below, choose a research mode, and download your completed Excel report.'
        '</p>',
        unsafe_allow_html=True,
    )


    # ── 4. Three-step guide cards ─────────────────────────────────────────────────
    _S = ("border:1px solid #E8E4DC;border-top:3px solid #B8960C;border-radius:8px;"
          "background:#FFFFFF;padding:28px;height:100%;box-sizing:border-box;"
          "box-shadow:0 1px 4px rgba(0,0,0,0.06);transition:box-shadow 0.2s;")
    sc1, sc2, sc3 = st.columns(3)
    with sc1:
        st.markdown(
            f'<div style="{_S}">'
            f'<div style="font-size:48px;line-height:1;font-family:Playfair Display,Georgia,serif;'
            f'font-weight:600;color:#B8960C;margin-bottom:12px;">1</div>'
            f'<div style="font-size:16px;font-weight:600;color:#1C2B3A;margin-bottom:8px;">'
            f'Enter Tickers</div>'
            f'<div style="font-size:14px;color:#6B7280;line-height:1.65;">'
            f'Type symbols like <code style="color:#B8960C;font-size:12px;'
            f'background:#FEF9EC;padding:1px 5px;border-radius:4px;">DIBS, FEAM, ACON</code> '
            f'or upload an Excel file with a "Ticker" column.</div>'
            f'</div>', unsafe_allow_html=True)
    with sc2:
        st.markdown(
            f'<div style="{_S}">'
            f'<div style="font-size:48px;line-height:1;font-family:Playfair Display,Georgia,serif;'
            f'font-weight:600;color:#B8960C;margin-bottom:12px;">2</div>'
            f'<div style="font-size:16px;font-weight:600;color:#1C2B3A;margin-bottom:8px;">'
            f'Run Research</div>'
            f'<div style="font-size:14px;color:#6B7280;line-height:1.65;">'
            f'Click Run below. Allow '
            f'<strong style="color:#1C2B3A;">30–45 seconds per ticker</strong>. '
            f'Keep this tab open — results appear automatically.</div>'
            f'</div>', unsafe_allow_html=True)
    with sc3:
        st.markdown(
            f'<div style="{_S}">'
            f'<div style="font-size:48px;line-height:1;font-family:Playfair Display,Georgia,serif;'
            f'font-weight:600;color:#B8960C;margin-bottom:12px;">3</div>'
            f'<div style="font-size:16px;font-weight:600;color:#1C2B3A;margin-bottom:8px;">'
            f'Download Report</div>'
            f'<div style="font-size:14px;color:#6B7280;line-height:1.65;">'
            f'Export as a formatted Excel file with '
            f'<strong style="color:#1C2B3A;">3 tabs</strong>: '
            f'Full Output, Contacts Only, and Financials Only.</div>'
            f'</div>', unsafe_allow_html=True)

    st.markdown("<hr style='margin:1rem 0;'>", unsafe_allow_html=True)


    # ── 5. Input area ─────────────────────────────────────────────────────────────
    left, right = st.columns([3, 2], gap="large")

    with left:
        # ── 5a. Research Mode selector (button group) ─────────────────────────────
        st.markdown(
            '<div style="font-size:11px;color:#9CA3AF;text-transform:uppercase;'
            'letter-spacing:0.14em;margin-bottom:6px;font-family:DM Sans,sans-serif;">'
            'Research Mode</div>',
            unsafe_allow_html=True,
        )
        _cur = st.session_state.research_mode
        mc1, mc2, mc3, _ = st.columns([0.9, 1.7, 1.8, 1.6])
        with mc1:
            if st.button(
                "Both",
                key="mode_both",
                type="primary" if _cur == "Both" else "secondary",
                use_container_width=True,
            ):
                st.session_state.research_mode = "Both"
                st.rerun()
        with mc2:
            if st.button(
                "Financial Only",
                key="mode_fin",
                type="primary" if _cur == "Financial Analysis Only" else "secondary",
                use_container_width=True,
            ):
                st.session_state.research_mode = "Financial Analysis Only"
                st.rerun()
        with mc3:
            if st.button(
                "Person Lookup Only",
                key="mode_person",
                type="primary" if _cur == "Person Lookup Only" else "secondary",
                use_container_width=True,
            ):
                st.session_state.research_mode = "Person Lookup Only"
                st.rerun()
        mode = st.session_state.research_mode

        _mode_desc = {
            "Both":                   "Full research — financials + CEO, CFO & IR contacts (default)",
            "Financial Analysis Only":"Price · Market Cap · Cash · Volume  —  fast, ~5–10s per ticker",
            "Person Lookup Only":     "CEO · CFO · IR emails & phones only  —  no financial data",
        }
        st.markdown(
            f'<div style="font-size:12px;color:#9CA3AF;margin:4px 0 10px;'
            f'font-style:italic;font-family:DM Sans,sans-serif;">'
            f'{_mode_desc.get(mode,"")}</div>',
            unsafe_allow_html=True,
        )

        # ── 5b. Quick-fill buttons + SPAC checkbox ────────────────────────────────
        qc1, qc2, qc3, qc_spac = st.columns([1.1, 1.3, 0.8, 3])
        with qc1:
            if st.button("US samples", key="ex_us", help="Load sample US tickers: DIBS, FEAM, ACON, AYTU, ECOR"):
                st.session_state.ticker_area = "DIBS, FEAM, ACON, AYTU, ECOR"
                st.rerun()
        with qc2:
            if st.button("CA samples", key="ex_ca", help="Load sample Canadian tickers: AUUA-CA, ARG-CA, ENW-CA"):
                st.session_state.ticker_area = "AUUA-CA, ARG-CA, ENW-CA"
                st.rerun()
        with qc3:
            if st.session_state.get("ticker_area") and st.button("Clear", key="clr"):
                st.session_state.ticker_area = ""
                st.rerun()
        with qc_spac:
            is_spac = st.checkbox(
                "SPAC mode",
                value=st.session_state.spac_mode,
                key="spac_checkbox",
                help="Enable when looking up SPACs (blank-check acquisition companies). "
                     "SPACs usually have no website, so the normal domain-based lookup fails. "
                     "This mode uses name-only searches as a fallback.",
            )
            st.session_state.spac_mode = is_spac

        # ── 5c. Ticker text area ──────────────────────────────────────────────────
        st.text_area(
            "TICKERS",
            placeholder=(
                "Type or paste tickers here — e.g. DIBS, FEAM, ACON, ATGN"
                " — commas, spaces, or one per line all work"
            ),
            height=96,
            help="Enter stock tickers separated by commas, spaces, or newlines. "
                 "Supports US (NYSE/NASDAQ/OTC) and Canadian (−CA suffix) tickers.",
            key="ticker_area",
        )
        ticker_raw = st.session_state.get("ticker_area", "")
        tickers = _parse_tickers(ticker_raw or "")

        # ── 5d. Live ticker chip display ──────────────────────────────────────────
        if tickers:
            chips = " ".join(
                f'<code style="background:#FFFFFF;border:1px solid #E8E4DC;border-radius:4px;'
                f'padding:2px 8px;font-size:12px;color:#1C2B3A;margin:2px;'
                f'font-family:JetBrains Mono,monospace;'
                f'box-shadow:0 1px 2px rgba(0,0,0,0.05);">{t}</code>'
                for t in tickers[:20]
            )
            more = (f'<a style="color:#B8960C;font-size:12px;text-decoration:none;cursor:default;">'
                    f'+{len(tickers)-20} more</a>'
                    if len(tickers) > 20 else "")
            st.markdown(
                f'<div style="margin:6px 0 8px;">'
                f'<span style="font-size:12px;color:#B8960C;font-weight:600;'
                f'font-family:DM Sans,sans-serif;">'
                f'{len(tickers)} ticker{"s" if len(tickers)!=1 else ""} detected'
                f'</span><span style="color:#E8E4DC;font-size:12px;"> — </span>'
                f'{chips}{more}</div>',
                unsafe_allow_html=True,
            )

        # ── 5e. Estimated run time ────────────────────────────────────────────────
        if tickers:
            secs_per = 10 if mode == "Financial Analysis Only" else 40
            est_lo = max(1, len(tickers) * secs_per // 60)
            est_hi = max(2, len(tickers) * (secs_per + 15) // 60)
            est_str = (f"~{len(tickers) * secs_per}s" if len(tickers) * secs_per < 90
                       else f"~{est_lo}–{est_hi} min")
            st.markdown(
                f'<div style="font-size:12px;color:#9CA3AF;margin-bottom:10px;'
                f'font-style:italic;font-family:DM Sans,sans-serif;">'
                f'Est. time: <span style="color:#6B7280;">{est_str}</span> · keep this tab open'
                f'</div>',
                unsafe_allow_html=True,
            )

        # ── 5f. Run Research button ───────────────────────────────────────────────
        run_ph = st.empty()
        run = run_ph.button(
            f"▶  Run Research"
            + (f"  —  {len(tickers)} Ticker{'s' if len(tickers)!=1 else ''}" if tickers else ""),
            type="primary",
            use_container_width=True,
            disabled=not tickers,
        )

    with right:
        # ── 5g. File uploader ─────────────────────────────────────────────────────
        uploaded = st.file_uploader(
            "UPLOAD TICKER LIST",
            type=["xlsx", "csv", "txt"],
            help=(
                'Upload an Excel, CSV, or TXT file.\n'
                'Must have a column labeled "Ticker" or "Symbol".\n'
                'Example: the Planet MicroCap spreadsheet.'
            ),
        )
        if uploaded:
            file_tickers = _tickers_from_file(uploaded)
            if file_tickers:
                st.session_state.ticker_area = "\n".join(file_tickers)
                ticker_raw = st.session_state.ticker_area
                tickers = _parse_tickers(ticker_raw or "")
                st.success(f"✓ {len(file_tickers)} tickers loaded from **{uploaded.name}** — ready to run")
            else:
                st.error(
                    'Could not find tickers in this file. '
                    'Make sure it has a column labeled "Ticker" or "Symbol".'
                )

    st.markdown("<hr style='margin:1rem 0;'>", unsafe_allow_html=True)


    # ══════════════════════════════════════════════════════════════════════════════
    # ── 6. Processing ─────────────────────────────────────────────────────────────
    # ══════════════════════════════════════════════════════════════════════════════

    if run and tickers:
        st.session_state.results      = []
        st.session_state.results_mode = mode
        st.session_state.running      = True

        run_ph.warning(
            "⟳  Running — do not close or refresh this tab while research is in progress.",
            icon="⏳",
        )

        secs_per  = 10 if mode == "Financial Analysis Only" else 40
        est_lo_m  = max(1, len(tickers) * secs_per // 60)
        est_hi_m  = max(2, len(tickers) * (secs_per + 15) // 60)
        est_banner = (f"~{len(tickers) * secs_per} seconds" if len(tickers) * secs_per < 90
                      else f"{est_lo_m}–{est_hi_m} minutes")
        st.info(
            f"Research started — processing **{len(tickers)}** "
            f"ticker{'s' if len(tickers)!=1 else ''}. "
            f"Estimated time: **{est_banner}**. "
            f"Results will appear below when complete.",
            icon="📊",
        )

        prog_ph   = st.empty()
        step_ph   = st.empty()
        log_ph    = st.empty()
        log_lines: list[str] = []

        def log(msg: str, level: str = "dim"):
            friendly = msg
            if "rate" in msg.lower() and ("limit" in msg.lower() or "429" in msg):
                friendly = "  ⏸ Rate limit reached — the tool is pausing before continuing automatically"
            elif ("connect" in msg.lower() or "timeout" in msg.lower() or "proxy" in msg.lower()):
                if "salesql" in msg.lower():
                    friendly = "  ⚠ Couldn't reach SalesQL — check your internet connection"
                elif "bouncer" in msg.lower():
                    friendly = "  ⚠ Couldn't reach Bouncer — email verification skipped for this ticker"
            cls   = {"ok":"ok","err":"err","warn":"warn","info":"info","dim":"dim"}.get(level,"dim")
            clean = friendly.replace("<","&lt;").replace(">","&gt;")
            log_lines.append(f'<span class="{cls}">{clean}</span>')
            log_ph.markdown(
                '<div class="run-log">' + "<br>".join(log_lines[-25:]) + "</div>",
                unsafe_allow_html=True,
            )

        def prog(step_text: str):
            step_ph.markdown(
                f'<div style="font-family:JetBrains Mono,monospace;font-size:0.74rem;'
                f'color:#C9A84C;margin:-0.2rem 0 0.3rem;letter-spacing:0.04em;">'
                f'⟳  {step_text}</div>',
                unsafe_allow_html=True,
            )

        t0   = time.time()
        rows: list[dict] = []
        n    = len(tickers)

        for i, tkr in enumerate(tickers):
            elapsed = time.time() - t0
            avg     = elapsed / max(i, 1)
            rem     = int(avg * (n - i))
            rem_str = f"  ·  ~{rem//60}m {rem%60}s remaining" if i > 0 and rem > 5 else ""
            prog_ph.progress(i / n, text=f"Ticker {i+1} of {n}{rem_str}")

            try:
                row = process_ticker(
                    tkr,
                    log_fn        = log,
                    prog_fn       = prog,
                    skip_linkedin = skip_li,
                    delay         = delay,
                    mode          = mode,
                    run_bouncer   = run_bouncer,
                    is_spac       = is_spac,
                )
                rows.append(row if row else empty_row(tkr, "Error"))
            except Exception as exc:
                err = str(exc)
                if "rate" in err.lower() or "429" in err:
                    log(f"  [{tkr}]  Rate limit — resuming automatically", "warn")
                elif "connect" in err.lower() or "timeout" in err.lower():
                    log(f"  [{tkr}]  Connection issue — skipping. Check internet connection.", "err")
                else:
                    log(f"  [{tkr}]  Symbol not recognised — double-check the spelling", "err")
                rows.append(empty_row(tkr, "Error"))

        elapsed_total = time.time() - t0
        prog_ph.progress(1.0, text=f"✅  Complete — {len(rows)} ticker{'s' if len(rows)!=1 else ''} processed")
        step_ph.empty()
        st.session_state.results       = rows
        st.session_state.elapsed_total = elapsed_total
        st.session_state.running       = False


    # ══════════════════════════════════════════════════════════════════════════════
    # ── 7. Results ────────────────────────────────────────────────────────────────
    # ══════════════════════════════════════════════════════════════════════════════

    if st.session_state.results:
        rows     = st.session_state.results
        res_mode = st.session_state.get("results_mode", "Both")
        elapsed  = st.session_state.get("elapsed_total", 0)

        # Summary bar
        total_f, missing_f = _count_populated(rows)
        populated  = total_f - missing_f
        elapsed_str = (f"{int(elapsed)//60}m {int(elapsed)%60}s"
                       if elapsed >= 60 else f"{int(elapsed)}s")
        bc = {"Both":"#B8960C","Financial Analysis Only":"#1A5C8A","Person Lookup Only":"#15803D"}.get(res_mode,"#B8960C")
        contact_stats = (
            f"<span style='font-size:13px;color:#E8E4DC;'>|</span>"
            f"<span style='font-size:13px;color:#6B7280;'>"
            f"{populated} fields populated · {missing_f} missing</span>"
            if res_mode != "Financial Analysis Only" else ""
        )
        st.markdown(
            f'<div style="background:#FFFFFF;border:1px solid #E8E4DC;border-radius:8px;'
            f'padding:14px 24px;margin-bottom:16px;display:flex;align-items:center;'
            f'flex-wrap:wrap;gap:14px;box-shadow:0 1px 4px rgba(0,0,0,0.06);">'
            f'<span style="font-size:13px;font-weight:600;color:#15803D;">'
            f'✓ Research complete</span>'
            f'<span style="font-size:13px;color:#E8E4DC;">|</span>'
            f'<span style="font-size:13px;color:#1C2B3A;">'
            f'{len(rows)} {"company" if len(rows)==1 else "companies"}</span>'
            f'{contact_stats}'
            f'<span style="font-size:13px;color:#E8E4DC;">|</span>'
            f'<span style="font-size:13px;color:#9CA3AF;">Completed in {elapsed_str}</span>'
            f'<span style="margin-left:auto;background:{bc}18;border:1px solid {bc}55;'
            f'color:{bc};font-size:11px;font-weight:600;letter-spacing:0.1em;'
            f'text-transform:uppercase;padding:4px 12px;border-radius:20px;">{res_mode}</span>'
            f'</div>',
            unsafe_allow_html=True,
        )

        # Company cards
        for row in rows:
            st.markdown(_company_card_html(row, res_mode), unsafe_allow_html=True)

        # Export section
        CONTACT_COLS_E   = ["Ticker","Company Name","CEO","CFO","CEO EMAIL","CEO NUMBER",
                            "CFO EMAIL","CFO NUMBER","IR Email","IR Contact","IR Page"]
        FINANCIAL_COLS_E = ["Ticker","Company Name","Industry","Exchange",
                            "Stock Price (Most Recent)","Market Cap (Most Recent)",
                            "Cash (Latest K)","Cash (Latest Q)","1M Share Volume",
                            "1D $ Share Volume","Cash from Ops (Latest K)","Cash from Ops (Latest Q)"]

        if res_mode == "Financial Analysis Only":
            export_cols = FINANCIAL_COLS_E
            sheet_note  = "The Excel file includes a **Financials** sheet."
        elif res_mode == "Person Lookup Only":
            export_cols = CONTACT_COLS_E
            sheet_note  = "The Excel file includes a **Contacts** sheet."
        else:
            export_cols = COLUMN_ORDER
            sheet_note  = "The Excel file includes **3 tabs**: Full Output (all data), Contacts Only, and Financials Only."

        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown(
            f'<div style="font-size:14px;color:#6B7280;margin-bottom:16px;'
            f'font-family:DM Sans,sans-serif;">'
            f'Your report is ready. {sheet_note}</div>',
            unsafe_allow_html=True,
        )
        fname = f"Curvature_Research_{date.today()}"
        dl1, dl2, _dl3 = st.columns([2, 2, 4])
        with dl1:
            st.download_button(
                "⬇  Download Excel Report",
                data      = build_excel(rows, mode=res_mode),
                file_name = fname + ".xlsx",
                mime      = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with dl2:
            safe_cols = [c for c in export_cols if c in (rows[0] if rows else {})]
            st.download_button(
                "⬇  Download CSV",
                data      = pd.DataFrame(rows, columns=safe_cols).to_csv(index=False),
                file_name = fname + ".csv",
                mime      = "text/csv",
                use_container_width=True,
            )

        # Collapsible raw data view
        exp_lbl = {
            "Both":                   "View all 21 columns",
            "Financial Analysis Only":"View financial columns",
            "Person Lookup Only":     "View contact columns",
        }.get(res_mode, "View data")
        with st.expander(exp_lbl):
            vc = [c for c in export_cols if c in rows[0]] if rows else export_cols
            st.dataframe(pd.DataFrame(rows, columns=vc), use_container_width=True, hide_index=True)

    else:
        # ── 8. Empty state ────────────────────────────────────────────────────────
        st.markdown(
            f'<div style="text-align:center;padding:64px 32px 48px;">'
            f'<div style="margin-bottom:24px;display:flex;justify-content:center;">'
            f'{_LOGO_SVG.replace("height:38px", "height:56px")}'
            f'</div>'
            f'<div style="font-size:15px;color:#9CA3AF;margin-bottom:8px;'
            f'font-weight:500;font-family:DM Sans,sans-serif;">'
            f'Enter tickers above to run your first research batch'
            f'</div>'
            f'<div style="font-size:13px;color:#D1CEC8;font-family:DM Sans,sans-serif;">'
            f'Single or batch input · Excel, CSV, TXT upload · Color-coded Excel export'
            f'</div>'
            f'</div>',
            unsafe_allow_html=True,
        )


# ══════════════════════════════════════════════════════════════════════════════
# ── PAGE: SPAC Research ──────────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════

elif st.session_state.nav_page == "SPAC Research":

    # ── Load roster ──────────────────────────────────────────────────────────
    _ROSTER_DIR = Path(__file__).resolve().parent / "spac_data"
    _ROSTER_FILE = _ROSTER_DIR / "spac_roster.json"
    roster = load_roster(_ROSTER_FILE)

    if not roster:
        st.warning("No SPAC roster found. Place `spac_roster.json` in the `spac_data/` folder.")
    else:
        # ── Header ───────────────────────────────────────────────────────────
        st.markdown(
            '<p style="font-size:14px;color:#6B7280;line-height:1.75;margin-bottom:16px;'
            'font-family:DM Sans,sans-serif;">'
            'Browse and enrich the SPAC roster — find CEO/CFO contact information via '
            '<span style="color:#B8960C;">SalesQL + LinkedIn</span>. '
            'Filter by urgency, search by ticker or company, then run enrichment on selected SPACs.'
            '</p>',
            unsafe_allow_html=True,
        )

        # ── Summary metrics ──────────────────────────────────────────────────
        _total       = len(roster)
        _with_ceo    = sum(1 for d in roster.values() if d.get("ceo_email"))
        _with_cfo    = sum(1 for d in roster.values() if d.get("cfo_email"))
        _needs_work  = sum(1 for d in roster.values() if needs_enrichment(d))
        _urgent      = sum(1 for d in roster.values() if (d.get("days_remaining") or 9999) < 30)
        _near_term   = sum(1 for d in roster.values() if 30 <= (d.get("days_remaining") or 9999) <= 90)

        mc1, mc2, mc3, mc4, mc5 = st.columns(5)
        mc1.metric("Total SPACs", _total)
        mc2.metric("CEO Emails Found", f"{_with_ceo}/{_total}")
        mc3.metric("CFO Emails Found", f"{_with_cfo}/{_total}")
        mc4.metric("Need Enrichment", _needs_work)
        mc5.metric("Urgent (<30d)", _urgent)

        st.markdown("<hr style='margin:0.75rem 0;'>", unsafe_allow_html=True)

        # ── Filters + Search ─────────────────────────────────────────────────
        fc1, fc2, fc3, fc4, _fc5 = st.columns([1.2, 1.2, 1.2, 1.5, 3])
        with fc1:
            if st.button(
                "All", key="spac_f_all",
                type="primary" if st.session_state.spac_filter == "All" else "secondary",
                use_container_width=True,
            ):
                st.session_state.spac_filter = "All"
                st.rerun()
        with fc2:
            if st.button(
                f"Urgent ({_urgent + _near_term})", key="spac_f_urgent",
                type="primary" if st.session_state.spac_filter == "Urgent" else "secondary",
                use_container_width=True,
            ):
                st.session_state.spac_filter = "Urgent"
                st.rerun()
        with fc3:
            if st.button(
                f"Missing ({_needs_work})", key="spac_f_missing",
                type="primary" if st.session_state.spac_filter == "Missing" else "secondary",
                use_container_width=True,
            ):
                st.session_state.spac_filter = "Missing"
                st.rerun()
        with fc4:
            spac_search = st.text_input(
                "Search", value=st.session_state.spac_search,
                placeholder="Ticker or company…",
                label_visibility="collapsed", key="spac_search_input",
            )
            st.session_state.spac_search = spac_search

        # ── Filter the roster ────────────────────────────────────────────────
        filtered_tickers = list(roster.keys())

        # Apply urgency filter
        if st.session_state.spac_filter == "Urgent":
            filtered_tickers = [
                t for t in filtered_tickers
                if (roster[t].get("days_remaining") or 9999) <= 90
            ]
        elif st.session_state.spac_filter == "Missing":
            filtered_tickers = [
                t for t in filtered_tickers
                if needs_enrichment(roster[t])
            ]

        # Apply search filter
        if spac_search.strip():
            q = spac_search.strip().upper()
            filtered_tickers = [
                t for t in filtered_tickers
                if q in t.upper() or q in roster[t].get("company", "").upper()
            ]

        # Sort by days_remaining (most urgent first)
        filtered_tickers.sort(key=lambda t: roster[t].get("days_remaining") or 9999)

        st.markdown(
            f'<div style="font-size:12px;color:#9CA3AF;margin:4px 0 10px;'
            f'font-family:DM Sans,sans-serif;">'
            f'Showing {len(filtered_tickers)} of {_total} SPACs</div>',
            unsafe_allow_html=True,
        )

        # ── Roster table ─────────────────────────────────────────────────────
        if filtered_tickers:
            table_rows = []
            for ticker in filtered_tickers:
                d = roster[ticker]
                days = d.get("days_remaining")
                urgency = get_urgency(d)

                # Urgency badge
                _urg_colours = {
                    "URGENT":    ("🔴", "#DC2626"),
                    "Near-term": ("🟠", "#D97706"),
                    "Upcoming":  ("🟡", "#CA8A04"),
                    "Standard":  ("🟢", "#16A34A"),
                    "Unknown":   ("⚪", "#9CA3AF"),
                }
                _emoji, _ = _urg_colours.get(urgency, ("⚪", "#9CA3AF"))

                table_rows.append({
                    "Ticker":    ticker,
                    "Company":   d.get("company", ""),
                    "Deadline":  d.get("deadline", ""),
                    "Days":      days if days is not None else "—",
                    "Urgency":   f"{_emoji} {urgency}",
                    "CEO":       d.get("ceo_name", ""),
                    "CEO Email": d.get("ceo_email", "") or "—",
                    "CFO":       d.get("cfo_name", ""),
                    "CFO Email": d.get("cfo_email", "") or "—",
                    "Source":    d.get("source", ""),
                })

            df_spac = pd.DataFrame(table_rows)
            st.dataframe(
                df_spac,
                use_container_width=True,
                hide_index=True,
                height=min(len(table_rows) * 35 + 40, 600),
            )

        else:
            st.info("No SPACs match your current filters.")

        st.markdown("<hr style='margin:0.75rem 0;'>", unsafe_allow_html=True)

        # ── Enrichment controls ──────────────────────────────────────────────
        ec1, ec2, ec3 = st.columns([2, 2, 4])
        with ec1:
            enrich_scope = st.selectbox(
                "Enrich scope",
                ["Filtered list", "Urgent only (<90 days)", "All missing"],
                key="enrich_scope",
            )
        with ec2:
            st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
            run_enrich = st.button(
                "▶  Run Enrichment",
                key="run_spac_enrich",
                type="primary",
                use_container_width=True,
                disabled=st.session_state.spac_enriching,
            )

        if run_enrich and not st.session_state.spac_enriching:
            st.session_state.spac_enriching = True

            # Determine targets
            if enrich_scope == "Filtered list":
                targets = [t for t in filtered_tickers if needs_enrichment(roster[t])]
            elif enrich_scope == "Urgent only (<90 days)":
                targets = [
                    t for t in roster
                    if needs_enrichment(roster[t])
                    and (roster[t].get("days_remaining") or 9999) < 90
                ]
            else:
                targets = [t for t in roster if needs_enrichment(roster[t])]

            targets.sort(key=lambda t: roster[t].get("days_remaining") or 9999)

            if not targets:
                st.info("All SPACs in this scope already have complete contacts.")
                st.session_state.spac_enriching = False
            else:
                import queue as _queue

                _n_workers = 5  # parallel workers (max 8 safe for SalesQL)
                st.markdown(
                    f'<div style="font-size:13px;color:#6B7280;margin-bottom:8px;'
                    f'font-family:DM Sans,sans-serif;">'
                    f'Enriching {len(targets)} SPACs with {_n_workers} parallel workers…</div>',
                    unsafe_allow_html=True,
                )

                progress_bar = st.progress(0)
                log_box = st.empty()
                log_lines: list[str] = []

                # Thread-safe log queue — workers append here,
                # main thread flushes to Streamlit UI
                _log_q: _queue.Queue = _queue.Queue()

                def _thread_log(msg: str, level: str = "dim"):
                    """Called from worker threads — never touches Streamlit."""
                    _log_q.put((msg, level))

                def _flush_logs():
                    """Called from main thread only — drains queue into UI."""
                    flushed = False
                    while not _log_q.empty():
                        try:
                            msg, lvl = _log_q.get_nowait()
                            css = lvl if lvl in ("ok","warn","err","info","dim") else "dim"
                            log_lines.append(f'<span class="{css}">{msg}</span>')
                            flushed = True
                        except _queue.Empty:
                            break
                    if flushed:
                        visible = log_lines[-30:]
                        log_box.markdown(
                            '<div class="run-log">' + "<br>".join(visible) + '</div>',
                            unsafe_allow_html=True,
                        )

                reset_cache()  # fresh cache for this run

                def _enrich_one_ui(ticker):
                    """Enrich one ticker in a worker thread. NO Streamlit calls."""
                    try:
                        roster[ticker] = enrich_spac_ticker(
                            ticker=ticker,
                            spac_data=roster[ticker],
                            delay=0.8,
                            skip_linkedin=False,
                            log_fn=_thread_log,
                        )
                    except Exception as e:
                        _thread_log(f"✗ Error on {ticker}: {e}", "err")
                    save_roster(roster, _ROSTER_FILE)
                    return ticker

                done_count = 0
                with _cf.ThreadPoolExecutor(max_workers=_n_workers) as executor:
                    futures = {
                        executor.submit(_enrich_one_ui, t): t
                        for t in targets
                    }
                    for future in _cf.as_completed(futures):
                        ticker = futures[future]
                        try:
                            future.result()
                        except Exception as e:
                            log_lines.append(f'<span class="err">✗ Unexpected: {e}</span>')
                        done_count += 1
                        # All UI updates happen here — in the main thread
                        pct = done_count / len(targets)
                        progress_bar.progress(pct, text=f"{ticker} done ({done_count}/{len(targets)})")
                        _flush_logs()

                _flush_logs()  # drain any remaining
                progress_bar.progress(1.0, text="Complete")
                log_lines.append(
                    f'<span class="ok">✓ Enrichment complete — '
                    f'{len(targets)} SPACs processed with {_n_workers} workers</span>'
                )
                log_box.markdown(
                    '<div class="run-log">' + "<br>".join(log_lines[-30:]) + '</div>',
                    unsafe_allow_html=True,
                )
                st.session_state.spac_enriching = False
                st.rerun()

        # ── Excel download ───────────────────────────────────────────────────
        st.markdown("<hr style='margin:0.75rem 0;'>", unsafe_allow_html=True)

        from spac_enricher import export_excel as _spac_export_excel

        dl1, dl2, _dl3 = st.columns([2, 2, 4])
        with dl1:
            try:
                _xlsx_buf = io.BytesIO()
                _tmp_path = Path("/tmp/_spac_export.xlsx")
                _spac_export_excel(roster, _tmp_path)
                with open(_tmp_path, "rb") as _xf:
                    _xlsx_data = _xf.read()
                st.download_button(
                    "⬇  Download SPAC Excel",
                    data=_xlsx_data,
                    file_name=f"spac_results_{date.today().isoformat()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception as e:
                st.caption(f"Excel export unavailable: {e}")

        with dl2:
            # CSV fallback
            csv_rows = []
            for ticker in sorted(roster.keys(), key=lambda t: roster[t].get("days_remaining") or 9999):
                d = roster[ticker]
                csv_rows.append({
                    "Ticker": ticker,
                    "Company": d.get("company", ""),
                    "Deadline": d.get("deadline", ""),
                    "Days Remaining": d.get("days_remaining"),
                    "CEO Name": d.get("ceo_name", ""),
                    "CEO Email": d.get("ceo_email", ""),
                    "CEO Phone": d.get("ceo_phone", ""),
                    "CFO Name": d.get("cfo_name", ""),
                    "CFO Email": d.get("cfo_email", ""),
                    "CFO Phone": d.get("cfo_phone", ""),
                    "Source": d.get("source", ""),
                    "Notes": d.get("notes", ""),
                })
            st.download_button(
                "⬇  Download CSV",
                data=pd.DataFrame(csv_rows).to_csv(index=False),
                file_name=f"spac_results_{date.today().isoformat()}.csv",
                mime="text/csv",
                use_container_width=True,
            )


# ══════════════════════════════════════════════════════════════════════════════
# ── PAGE: Bulk Enrich ───────────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════

elif st.session_state.nav_page == "Bulk Enrich":

    st.markdown(
        '<p style="font-size:14px;color:#6B7280;line-height:1.75;margin-bottom:24px;'
        'font-family:DM Sans,sans-serif;">'
        'Upload a spreadsheet with <span style="color:#B8960C;">company names</span> and '
        '<span style="color:#B8960C;">executive names/titles</span> — the platform will '
        'find their emails and phone numbers via SalesQL + LinkedIn.</p>',
        unsafe_allow_html=True,
    )

    # ── Session state defaults ───────────────────────────────────────────────
    if "bulk_df" not in st.session_state:
        st.session_state.bulk_df = None
    if "bulk_col_map" not in st.session_state:
        st.session_state.bulk_col_map = {}
    if "bulk_results" not in st.session_state:
        st.session_state.bulk_results = None
    if "bulk_running" not in st.session_state:
        st.session_state.bulk_running = False

    # ── File upload ──────────────────────────────────────────────────────────
    uploaded = st.file_uploader(
        "Upload spreadsheet (.xlsx or .csv)",
        type=["xlsx", "csv"],
        key="bulk_upload",
    )

    if uploaded is not None:
        try:
            if uploaded.name.endswith(".csv"):
                df = pd.read_csv(uploaded)
            else:
                # Try to read — auto-detect header row
                df = pd.read_excel(uploaded, engine="openpyxl")
                # If first row looks like a title (all nulls after col 0), skip rows
                if df.shape[1] > 3 and df.iloc[0].isna().sum() > df.shape[1] * 0.6:
                    # Re-read skipping title rows — find the real header
                    for skip in range(1, 6):
                        df = pd.read_excel(uploaded, engine="openpyxl", header=skip)
                        if df.columns.dtype == object and not df.columns[0] is None:
                            break
            st.session_state.bulk_df = df
        except Exception as e:
            st.error(f"Could not read file: {e}")
            st.session_state.bulk_df = None

    df = st.session_state.bulk_df

    if df is not None:
        st.markdown(f"**{len(df)} rows  ·  {len(df.columns)} columns**")

        # ── Preview ──────────────────────────────────────────────────────────
        with st.expander("Preview uploaded data", expanded=False):
            st.dataframe(df.head(20), use_container_width=True)

        # ── Column mapping ───────────────────────────────────────────────────
        st.markdown("##### Map your columns")
        st.markdown(
            '<p style="font-size:13px;color:#6B7280;">Tell us which columns contain '
            'the company name, person name, and title. The rest is automatic.</p>',
            unsafe_allow_html=True,
        )

        all_cols = ["— (not mapped)"] + list(df.columns)

        # Auto-detect columns by common names
        def _auto_detect(candidates: list[str]) -> str:
            for c in df.columns:
                if str(c).strip().lower() in candidates:
                    return str(c)
            # Partial match
            for c in df.columns:
                cl = str(c).strip().lower()
                for cand in candidates:
                    if cand in cl or cl in cand:
                        return str(c)
            return "— (not mapped)"

        _def_company = _auto_detect(["company", "company name", "organization", "firm"])
        _def_name    = _auto_detect(["name", "full name", "executive name", "contact name", "person", "executive"])
        _def_title   = _auto_detect(["title", "role", "job title", "position"])
        _def_website = _auto_detect(["website", "url", "domain", "company website", "web"])
        _def_ticker  = _auto_detect(["ticker", "symbol", "stock"])

        mc1, mc2, mc3 = st.columns(3)
        with mc1:
            col_company = st.selectbox("Company", all_cols, index=all_cols.index(_def_company) if _def_company in all_cols else 0, key="bulk_map_company")
        with mc2:
            col_name = st.selectbox("Person Name", all_cols, index=all_cols.index(_def_name) if _def_name in all_cols else 0, key="bulk_map_name")
        with mc3:
            col_title = st.selectbox("Title / Role", all_cols, index=all_cols.index(_def_title) if _def_title in all_cols else 0, key="bulk_map_title")

        mc4, mc5, mc6 = st.columns(3)
        with mc4:
            col_website = st.selectbox("Website (optional)", all_cols, index=all_cols.index(_def_website) if _def_website in all_cols else 0, key="bulk_map_website")
        with mc5:
            col_ticker = st.selectbox("Ticker (optional)", all_cols, index=all_cols.index(_def_ticker) if _def_ticker in all_cols else 0, key="bulk_map_ticker")
        with mc6:
            col_linkedin = st.selectbox("LinkedIn URL (optional)", all_cols, index=all_cols.index(_auto_detect(["linkedin", "linkedin url", "linkedin_url", "li url"])) if _auto_detect(["linkedin", "linkedin url", "linkedin_url", "li url"]) in all_cols else 0, key="bulk_map_linkedin")

        _unmapped = "— (not mapped)"
        _has_required = col_company != _unmapped and col_name != _unmapped

        if not _has_required:
            st.warning("Please map at least **Company** and **Person Name** columns to proceed.")

        # ── Enrichment settings ──────────────────────────────────────────────
        st.markdown("---")
        ec1, ec2 = st.columns(2)
        with ec1:
            bulk_use_linkedin = st.checkbox("Enable LinkedIn search", value=True, key="bulk_linkedin")
        with ec2:
            bulk_verify_emails = st.checkbox("Verify emails (Bouncer)", value=True, key="bulk_verify")

        # ── Run enrichment ───────────────────────────────────────────────────
        if _has_required and st.button("🚀  Enrich Contacts", type="primary", use_container_width=True, key="bulk_run"):
            st.session_state.bulk_running = True
            _delay = 0.8

            progress_bar = st.progress(0)
            status_text  = st.empty()
            log_box      = st.empty()
            log_lines: list[str] = []

            def _blog(msg: str, level: str = "dim"):
                css = level if level in ("ok", "warn", "err", "info", "dim") else "dim"
                log_lines.append(f'<span class="{css}">{msg}</span>')
                visible = log_lines[-25:]
                log_box.markdown(
                    '<div class="run-log">' + "<br>".join(visible) + '</div>',
                    unsafe_allow_html=True,
                )

            results: list[dict] = []
            total = len(df)

            for idx, raw_row in df.iterrows():
                i = int(idx) if isinstance(idx, (int, float)) else results.__len__()
                pct = (i + 1) / total
                progress_bar.progress(pct)

                company  = str(raw_row.get(col_company, "")).strip() if col_company != _unmapped else ""
                name     = str(raw_row.get(col_name, "")).strip() if col_name != _unmapped else ""
                title    = str(raw_row.get(col_title, "")).strip() if col_title != _unmapped else ""
                website  = str(raw_row.get(col_website, "")).strip() if col_website != _unmapped else ""
                ticker   = str(raw_row.get(col_ticker, "")).strip() if col_ticker != _unmapped else ""
                li_url   = str(raw_row.get(col_linkedin, "")).strip() if col_linkedin != _unmapped else ""

                # Clean up nan values from pandas
                for _v in [company, name, title, website, ticker, li_url]:
                    if _v.lower() in ("nan", "none", ""):
                        pass  # handled below
                company = "" if company.lower() in ("nan", "none") else company
                name    = "" if name.lower() in ("nan", "none") else name
                title   = "" if title.lower() in ("nan", "none") else title
                website = "" if website.lower() in ("nan", "none") else website
                ticker  = "" if ticker.lower() in ("nan", "none") else ticker
                li_url  = "" if li_url.lower() in ("nan", "none") else li_url

                if not company or not name:
                    _blog(f"⚠  Row {i+1}: skipped (missing company or name)", "warn")
                    results.append({
                        "Row": i + 1, "Ticker": ticker, "Company": company,
                        "Name": name, "Title": title,
                        "Email": "", "Phone": "", "LinkedIn": li_url,
                        "Source": "skipped",
                    })
                    continue

                status_text.markdown(f"**{i+1}/{total}**  ·  {name} @ {company}")
                _blog(f"[{i+1}/{total}]  {name}  ·  {company}  ·  {title}", "info")

                enrichment: dict = {}
                source = ""

                try:
                    # ── Step 1: If LinkedIn URL is provided, enrich directly ──
                    if li_url and "linkedin.com/in/" in li_url:
                        _blog(f"   LinkedIn URL provided → SalesQL enriching …", "dim")
                        time.sleep(_delay)
                        enrichment = enrich_by_url(li_url)
                        if enrichment.get("best_email") or enrichment.get("phone"):
                            source = "salesql_linkedin"
                            _blog(f"  ✓ {enrichment.get('best_email','—')}  {enrichment.get('phone','—')}", "ok")

                    # ── Step 2: SalesQL name search ──────────────────────────
                    if not enrichment.get("best_email") and not enrichment.get("phone"):
                        _blog(f"   SalesQL name search …", "dim")
                        first, last = split_name(name)
                        time.sleep(_delay)
                        enrichment = search_by_name_with_variations(
                            first, last, name, company,
                            website=website,
                            is_spac=False,
                        )
                        if enrichment.get("best_email") or enrichment.get("phone"):
                            source = "salesql_name"
                            if not li_url:
                                li_url = enrichment.get("linkedin_url", "")
                            _blog(f"  ✓ {enrichment.get('best_email','—')}  {enrichment.get('phone','—')}", "ok")

                    # ── Step 3: LinkedIn search + SalesQL enrich ─────────────
                    if not enrichment.get("best_email") and not enrichment.get("phone") and bulk_use_linkedin:
                        _blog(f"   LinkedIn search → {name} …", "dim")
                        found_li = find_linkedin_url(name, company, title, sleep_range=(_delay, _delay + 0.3))
                        if found_li:
                            li_url = found_li
                            _blog(f"   LinkedIn: {found_li} → enriching …", "dim")
                            time.sleep(_delay)
                            enrichment = enrich_by_url(found_li)
                            if enrichment.get("best_email") or enrichment.get("phone"):
                                source = "linkedin_enrich"
                                _blog(f"  ✓ {enrichment.get('best_email','—')}  {enrichment.get('phone','—')}", "ok")

                    # ── Step 4: Bouncer verification ─────────────────────────
                    best_email = enrichment.get("best_email", "")
                    if best_email and bulk_verify_emails:
                        from lookup.bouncer_verifier import verify_email, apply_flag
                        _blog(f"   Bouncer: verifying {best_email} …", "dim")
                        time.sleep(0.3)
                        vresult = verify_email(best_email)
                        best_email = apply_flag(best_email, vresult)
                        vstatus = vresult.get("status", "unknown")
                        if vstatus == "deliverable":
                            _blog(f"   ✓ Bouncer: deliverable", "ok")
                        elif vstatus in ("undeliverable", "unknown"):
                            _blog(f"   ✗ Bouncer: {vstatus} — email removed", "warn")
                        else:
                            _blog(f"   ⚠ Bouncer: {vstatus}", "warn")

                    if not source:
                        source = "not_found"
                        _blog(f"  ⚠ No contact found", "warn")

                except Exception as e:
                    _blog(f"  ✗ Error: {e}", "err")
                    source = "error"
                    best_email = ""

                results.append({
                    "Row": i + 1,
                    "Ticker": ticker,
                    "Company": company,
                    "Name": name,
                    "Title": title,
                    "Email": best_email if isinstance(best_email, str) else enrichment.get("best_email", ""),
                    "Phone": enrichment.get("phone", "") or "",
                    "LinkedIn": li_url or enrichment.get("linkedin_url", "") or "",
                    "Source": source,
                })

            progress_bar.progress(1.0)
            status_text.markdown("**✅ Enrichment complete**")

            result_df = pd.DataFrame(results)
            st.session_state.bulk_results = result_df
            st.session_state.bulk_running = False

            # Stats
            found = len([r for r in results if r.get("Email") or r.get("Phone")])
            _blog(f"\n✅  Done — {found}/{total} contacts enriched", "ok")

        # ── Show results ─────────────────────────────────────────────────────
        if st.session_state.bulk_results is not None and not st.session_state.bulk_running:
            rdf = st.session_state.bulk_results
            st.markdown("---")
            st.markdown("##### Enriched Results")

            # Metrics
            total_r  = len(rdf)
            found_r  = len(rdf[(rdf["Email"].str.len() > 0) | (rdf["Phone"].str.len() > 0)])
            email_r  = len(rdf[rdf["Email"].str.len() > 0])
            phone_r  = len(rdf[rdf["Phone"].str.len() > 0])
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Total", total_r)
            m2.metric("Found", found_r)
            m3.metric("Emails", email_r)
            m4.metric("Phones", phone_r)

            st.dataframe(
                rdf,
                use_container_width=True,
                column_config={
                    "LinkedIn": st.column_config.LinkColumn("LinkedIn"),
                    "Email": st.column_config.TextColumn("Email"),
                },
            )

            # ── Download buttons ─────────────────────────────────────────────
            dl1, dl2 = st.columns(2)
            with dl1:
                csv_data = rdf.to_csv(index=False)
                st.download_button(
                    "⬇  Download CSV",
                    data=csv_data,
                    file_name=f"enriched_contacts_{date.today().isoformat()}.csv",
                    mime="text/csv",
                    use_container_width=True,
                )
            with dl2:
                # Excel with formatting
                buf = io.BytesIO()
                with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                    rdf.to_excel(writer, index=False, sheet_name="Enriched Contacts")
                    ws = writer.sheets["Enriched Contacts"]
                    from openpyxl.styles import Font, PatternFill, Alignment
                    header_font = Font(bold=True, color="FFFFFF", size=11)
                    header_fill = PatternFill(start_color="1C2B3A", end_color="1C2B3A", fill_type="solid")
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")
                    for col in ws.columns:
                        max_len = max(len(str(c.value or "")) for c in col)
                        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
                buf.seek(0)
                st.download_button(
                    "⬇  Download Excel",
                    data=buf.getvalue(),
                    file_name=f"enriched_contacts_{date.today().isoformat()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
