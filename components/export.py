"""
components/export.py
Excel + CSV export for Curvature Securities IB Intelligence Platform.

Produces a 3-sheet .xlsx:
  Sheet 1 — Full Output      (all 21 columns)
  Sheet 2 — Contacts Only    (Ticker + all contact fields)
  Sheet 3 — Financials Only  (Ticker + Company + all financial fields)
"""

from __future__ import annotations

import io
from datetime import date
from typing import Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

import pandas as pd

# ── Column groups ─────────────────────────────────────────────────────────────
COLUMN_ORDER = [
    "Company Name", "Ticker", "Industry", "Exchange",
    "Stock Price (Most Recent)", "Market Cap (Most Recent)",
    "Cash (Latest K)", "Cash (Latest Q)",
    "1M Share Volume", "1D $ Share Volume",
    "Cash from Ops (Latest K)", "Cash from Ops (Latest Q)",
    "CEO", "CFO",
    "CEO EMAIL", "CEO NUMBER",
    "CFO EMAIL", "CFO NUMBER",
    "IR Email", "IR Contact", "IR Page",
]

CONTACT_COLS = [
    "Ticker", "Company Name",
    "CEO", "CFO",
    "CEO EMAIL", "CEO NUMBER",
    "CFO EMAIL", "CFO NUMBER",
    "IR Email", "IR Contact", "IR Page",
]

FINANCIAL_COLS = [
    "Ticker", "Company Name", "Industry", "Exchange",
    "Stock Price (Most Recent)", "Market Cap (Most Recent)",
    "Cash (Latest K)", "Cash (Latest Q)",
    "1M Share Volume", "1D $ Share Volume",
    "Cash from Ops (Latest K)", "Cash from Ops (Latest Q)",
]

# ── Palette ───────────────────────────────────────────────────────────────────
_GOLD_HEX     = "C9A84C"
_DARK_BG      = "0D0F14"
_CARD_BG      = "111318"
_ROW_A        = "111318"
_ROW_B        = "161A22"
_TEXT_MAIN    = "F0EDE8"
_TEXT_MUTED   = "4A4D56"
_GREEN        = "2ECC71"
_ORANGE       = "F39C12"
_RED          = "E74C3C"

# ── Style builders ────────────────────────────────────────────────────────────

def _hdr_style():
    return (
        Font(bold=True, color=_DARK_BG, size=9, name="DM Sans"),
        PatternFill("solid", fgColor=_GOLD_HEX),
        Alignment(horizontal="center", vertical="center", wrap_text=True),
    )


def _cell_style(row_idx: int):
    fill_hex = _ROW_A if row_idx % 2 == 0 else _ROW_B
    return (
        Font(color=_TEXT_MAIN, size=9, name="Calibri"),
        PatternFill("solid", fgColor=fill_hex),
        Alignment(vertical="top", wrap_text=True),
    )


def _contact_font_and_fill(value: str, col: str, row_idx: int) -> tuple:
    """Return (Font, PatternFill) for contact cells with color-coding."""
    base_fill = PatternFill("solid", fgColor=_ROW_A if row_idx % 2 == 0 else _ROW_B)
    v = str(value or "").strip().lower()
    is_blank = v in {"", "not found", "not on linkedin", "not on sql", "n/a", "—"}

    if "EMAIL" in col:
        if is_blank:
            return Font(color=_TEXT_MUTED, size=9, italic=True, name="Calibri"), base_fill
        if "(no work provided)" in v:
            return Font(color=_ORANGE, size=9, name="JetBrains Mono"), base_fill
        return Font(color=_TEXT_MAIN, size=9, name="JetBrains Mono"), base_fill

    if "NUMBER" in col:
        if is_blank:
            return Font(color=_TEXT_MUTED, size=9, italic=True, name="Calibri"), base_fill
        if v.startswith("work"):
            return Font(color=_GREEN, size=9, name="JetBrains Mono"), base_fill
        return Font(color=_TEXT_MAIN, size=9, name="JetBrains Mono"), base_fill

    if is_blank:
        return Font(color=_TEXT_MUTED, size=9, italic=True, name="Calibri"), base_fill
    return Font(color=_TEXT_MAIN, size=9, name="Calibri"), base_fill


# ── Column widths ─────────────────────────────────────────────────────────────
_COL_WIDTHS = {
    "Company Name":              28,
    "Ticker":                    10,
    "Industry":                  20,
    "Exchange":                   9,
    "Stock Price (Most Recent)": 14,
    "Market Cap (Most Recent)":  16,
    "Cash (Latest K)":           14,
    "Cash (Latest Q)":           14,
    "1M Share Volume":           15,
    "1D $ Share Volume":         16,
    "Cash from Ops (Latest K)":  18,
    "Cash from Ops (Latest Q)":  18,
    "CEO":                       26,
    "CFO":                       26,
    "CEO EMAIL":                 34,
    "CEO NUMBER":                22,
    "CFO EMAIL":                 34,
    "CFO NUMBER":                22,
    "IR Email":                  30,
    "IR Contact":                28,
    "IR Page":                   38,
}


# ── Write helpers ─────────────────────────────────────────────────────────────

def _write_sheet(ws, rows: list[dict], cols: list[str]) -> None:
    """Write header + data rows to a worksheet with full styling."""
    h_font, h_fill, h_align = _hdr_style()

    # Header row
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font      = h_font
        cell.fill      = h_fill
        cell.alignment = h_align
    ws.row_dimensions[1].height = 22

    # Data rows
    for ri, row in enumerate(rows, 2):
        d_font, d_fill, d_align = _cell_style(ri)
        for ci, col in enumerate(cols, 1):
            val  = row.get(col, "")
            cell = ws.cell(row=ri, column=ci, value=val)

            # Contact columns get special color treatment
            if col in {"CEO EMAIL", "CFO EMAIL", "IR Email",
                        "CEO NUMBER", "CFO NUMBER"}:
                cf, cf_fill = _contact_font_and_fill(val, col, ri)
                cell.font  = cf
                cell.fill  = cf_fill
            else:
                cell.font  = d_font
                cell.fill  = d_fill

            cell.alignment = d_align

    # Column widths
    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = _COL_WIDTHS.get(col, 16)

    # Freeze header
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    # Sheet tab color (gold)
    ws.sheet_properties.tabColor = _GOLD_HEX


def rows_to_xlsx(rows: list[dict], filename_hint: str = "") -> bytes:
    """
    Build a 3-sheet .xlsx from enrichment rows.
    Returns raw bytes suitable for st.download_button.
    """
    if not HAS_OPENPYXL:
        # Fallback — single-sheet via pandas
        df  = pd.DataFrame(rows, columns=COLUMN_ORDER)
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as wr:
            df.to_excel(wr, index=False, sheet_name="Full Output")
        return buf.getvalue()

    wb = openpyxl.Workbook()

    # ── Sheet 1: Full Output ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Full Output"
    ws1.sheet_view.showGridLines = False
    _write_sheet(ws1, rows, COLUMN_ORDER)

    # ── Sheet 2: Contacts Only ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Contacts Only")
    ws2.sheet_view.showGridLines = False
    _write_sheet(ws2, rows, CONTACT_COLS)

    # ── Sheet 3: Financials Only ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Financials Only")
    ws3.sheet_view.showGridLines = False
    _write_sheet(ws3, rows, FINANCIAL_COLS)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def rows_to_csv(rows: list[dict]) -> str:
    """Return CSV string of all 21 columns."""
    df = pd.DataFrame(rows, columns=COLUMN_ORDER)
    return df.to_csv(index=False)


def export_filename(ext: str = "xlsx") -> str:
    """e.g. Curvature_Research_2025-01-15.xlsx"""
    return f"Curvature_Research_{date.today()}.{ext}"
